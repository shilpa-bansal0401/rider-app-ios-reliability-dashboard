#!/usr/bin/env python3
"""
Fetch BigQuery iOS user counts + Sentry hang/crash data for the current month,
then generate both outputs in a single run:

  1. sentry_data.xlsx              — Dashboard + Consolidation Excel workbook
  2. consolidation_report/index.html — HTML AQS consolidation report (all brands)

Usage:
    export SENTRY_AUTH_TOKEN=sntryu_...
    python3 fetch_sentry_data.py
"""

import urllib.request
import urllib.parse
import json
import os
import datetime
import sys
import time
import warnings
import calendar

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment
except ImportError:
    raise SystemExit("ERROR: run  pip install openpyxl  first.")

try:
    warnings.filterwarnings("ignore")
    from google.cloud import bigquery as bq_client
    BQ_AVAILABLE = True
except ImportError:
    BQ_AVAILABLE = False

try:
    import gspread
    from google.auth import default as _gauth_default
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False

SPREADSHEET_ID = "1IgBzRjH8XdStzD3L5-IyIg2A2_T4IjheChuhIm6drrc"
WORKSHEET_GID  = 138720290   # gid= from the URL fragment

TOKEN   = os.environ.get("SENTRY_AUTH_TOKEN", "")
ORG     = "delivery-hero-pm"
PROJECT = "4506937839976448"

if not TOKEN:
    raise SystemExit("ERROR: set SENTRY_AUTH_TOKEN environment variable first.")

TODAY = datetime.date.today()

if TODAY.day < 3:
    # On day 1-2 of the month, report on the previous full month
    first_of_current = TODAY.replace(day=1)
    END_DATE   = first_of_current - datetime.timedelta(days=1)  # last day of prev month
    START_DATE = END_DATE.replace(day=1)                         # 1st of prev month
else:
    # From day 3 onwards, report on the current month up to yesterday
    END_DATE   = TODAY - datetime.timedelta(days=1)
    START_DATE = TODAY.replace(day=1)

START    = START_DATE.strftime("%Y-%m-%dT00:00:00.000")
END      = END_DATE.strftime("%Y-%m-%dT23:59:59.999")
BQ_START = START_DATE.strftime("%Y-%m-%d")
BQ_END   = END_DATE.strftime("%Y-%m-%d")

# On day 3 only: also show the previous full month as a second HTML tab
SHOW_PREV_TAB = TODAY.day == 3
if SHOW_PREV_TAB:
    _prev_last      = TODAY.replace(day=1) - datetime.timedelta(days=1)
    PREV_START_DATE = _prev_last.replace(day=1)
    PREV_END_DATE   = _prev_last
    PREV_START    = PREV_START_DATE.strftime("%Y-%m-%dT00:00:00.000")
    PREV_END      = PREV_END_DATE.strftime("%Y-%m-%dT23:59:59.999")
    PREV_BQ_START = PREV_START_DATE.strftime("%Y-%m-%d")
    PREV_BQ_END   = PREV_END_DATE.strftime("%Y-%m-%d")
else:
    PREV_START_DATE = PREV_END_DATE = PREV_START = PREV_END = PREV_BQ_START = PREV_BQ_END = None

# Fixed 30-day window for per-release metrics — gradual rollout means a release
# may have started shipping weeks ago, so we always look back 30 days.
REL_END_DATE   = TODAY - datetime.timedelta(days=1)
REL_START_DATE = TODAY - datetime.timedelta(days=30)
REL_START    = REL_START_DATE.strftime("%Y-%m-%dT00:00:00.000")
REL_END      = REL_END_DATE.strftime("%Y-%m-%dT23:59:59.999")
REL_BQ_START = REL_START_DATE.strftime("%Y-%m-%d")
REL_BQ_END   = REL_END_DATE.strftime("%Y-%m-%d")

HANGS_QUERY   = '!user.id:*-*-*-*-* app.in_foreground:True "*App hang* detected*"'
CRASHES_QUERY = "level:fatal handled:no !stack.package:*gpsmaster* !stack.package:*GPSTraveller* !user.id:*-*-*-*-* !issue:RIDER-APP-IOS-Z7 !issue:RIDER-APP-IOS-1BM !issue:RIDER-APP-IOS-3DVQ"

HEADERS = {"Authorization": f"Bearer {TOKEN}"}

# Firebase Performance — brand display name → BigQuery table suffix
FIREBASE_BRANDS = [
    ("Foodpanda",     "com_logistics_rider_foodpanda_IOS"),
    ("Foodora",       "com_logistics_rider_foodora_IOS"),
    ("Talabat",       "com_logistics_rider_talabat_IOS"),
    ("pedidosya",     "com_logistics_rider_pedidosya_IOS"),
    ("HungerStation", "com_logistics_rider_hungerstation_IOS"),
    ("Yemeksepeti",   "com_logistics_rider_yemeksepeti_IOS"),
    ("Glovo",         "com_logistics_rider_glovo_IOS"),
    ("Woowa",         "com_logistics_rider_woowabros_IOS"),
    ("efood",         "com_logistics_rider_efood_IOS"),
    ("Foody",         "com_logistics_rider_foody_IOS"),
]

# Derived from FIREBASE_BRANDS for use in Firebase fetch queries
FIREBASE_TABLES = [t for _, t in FIREBASE_BRANDS]

# Pinned release versions tracked for version-based AQS performance scoring
PINNED_RELEASES = [
    {"version": "4.2634.1", "dist":  "1049",           "asti": 3.78,  "stti": 1.08},
    {"version": "4.2635.1", "dist":  "1052",           "asti": 3.78,  "stti": 1.08},
]


def rel_filter_str(rel):
    """Build a Sentry query filter for a PINNED_RELEASES entry (dist: or release:)."""
    dists = rel.get("dists")
    d     = rel.get("dist")
    v     = rel.get("version")
    if dists:
        return "(" + " OR ".join(f"dist:{dd}" for dd in dists) + ")"
    if d:
        return f"dist:{d}"
    return f"release:{v}"


# ── Brand definitions (shared by Excel and HTML outputs) ──────────────────────

BRANDS = [
    {"name": "Foodpanda",     "bq_key": "foodpanda",     "sentry_key": "foodpanda",     "app_size": 77.8, "asti": 3.93, "stti": 1.20, "riders": 78203},
    {"name": "Foodora",       "bq_key": "foodora",       "sentry_key": "foodora",       "app_size": 83.2, "asti": 3.93, "stti": 1.20, "riders": 14946},
    {"name": "Talabat",       "bq_key": "talabat",       "sentry_key": "talabat",       "app_size": 77.8, "asti": 3.93, "stti": 1.20, "riders": 16988},
    {"name": "pedidosya",     "bq_key": "pedidosya",     "sentry_key": "pedidosya",     "app_size": 68.4, "asti": 3.93, "stti": 1.20, "riders": 20641},
    {"name": "HungerStation", "bq_key": "hungerstation", "sentry_key": "hungerstation", "app_size": 78.3, "asti": 3.93, "stti": 1.20, "riders": 10630},
    {"name": "Yemeksepeti",   "bq_key": "yemeksepeti",   "sentry_key": "yemeksepeti",   "app_size": 78.3, "asti": 3.93, "stti": 1.20, "riders": 2186},
    {"name": "Glovo",         "bq_key": "glovo",         "sentry_key": "glovo",         "app_size": 77.9, "asti": 3.93, "stti": 1.20, "riders": 46896},
    {"name": "Woowa",         "bq_key": "woowabros",     "sentry_key": "woowa",         "app_size": 78.0, "asti": 3.93, "stti": 1.20, "riders": 731},
    {"name": "efood",         "bq_key": "efood",         "sentry_key": "efood",         "app_size": 68.1, "asti": 3.93, "stti": 1.20, "riders": 5258},
    {"name": "Foody",         "bq_key": "foody",         "sentry_key": "foody",         "app_size": 68.1, "asti": 3.93, "stti": 1.20, "riders": 729},
]

WEIGHTS      = [0.3966, 0.0758, 0.0861, 0.1047, 0.0539, 0.0111, 0.2378, 0.0037, 0.0267, 0.0037]
RIDER_COUNTS = [78203, 14946, 16988, 20641, 10630, 2186, 46896, 731, 5258, 729]

# ── AQS config ─────────────────────────────────────────────────────────────────

AQS_CONFIG = {
    "cfu":      {"baseline": 99.7, "target": 99.9, "weight": 35},
    "hang":     {"baseline": 99.7, "target": 99.9, "weight": 25},
    "app_size": {"baseline": 75,   "target": 60,   "weight": 2},
    "asti":     {"baseline": 4,    "target": 2,    "weight": 10},
    "stti":     {"baseline": 1.5,  "target": 0.5,  "weight": 10},
    "frozen":   {"baseline": 3,    "target": 1,    "weight": 13},
    "skipped":  {"baseline": 2,    "target": 1,    "weight": 5},
}

# ── Excel style constants ──────────────────────────────────────────────────────

BQ_HEADER_FILL    = PatternFill("solid", fgColor="BDD7EE")  # pastel sky blue
CRASH_HEADER_FILL = PatternFill("solid", fgColor="FFADAD")  # pastel rose
HANG_HEADER_FILL  = PatternFill("solid", fgColor="C9B8F0")  # pastel lavender
SETTINGS_FILL     = PatternFill("solid", fgColor="C6EFCE")  # pastel mint green
DASH_HEADER_FILL  = PatternFill("solid", fgColor="B8D4E8")  # pastel steel blue
EVEN_ROW_FILL     = PatternFill("solid", fgColor="EEF4FB")  # very light pastel blue
SUMMARY_LBL_FILL  = PatternFill("solid", fgColor="FDEBD0")  # pastel peach
SUMMARY_VAL_FILL  = PatternFill("solid", fgColor="D5F5E3")  # pastel sage green
BOLD = Font(bold=True)

_THIN   = Side(style="thin",   color="BFBFBF")
_MEDIUM = Side(style="medium", color="808080")
THIN_BOX   = Border(left=_THIN,   right=_THIN,   top=_THIN,   bottom=_THIN)
MEDIUM_BOX = Border(left=_MEDIUM, right=_MEDIUM, top=_MEDIUM, bottom=_MEDIUM)

# Raw data column positions — must match the SUMIFS formulas in the dashboard:
#   D =SUMIFS(V:V,  U:U, "*"&$B$10&"*", T:T,  C{row})  → BQ at T(20)-V(22)
#   E =SUMIFS(X:X,  Z:Z, "*"&$B$10&"*", Y:Y,  C{row})  → Crash at X(24)-Z(26)
#   K =SUMIFS(AB:AB,AD:AD,"*"&$B$10&"*",AC:AC,C{row})  → Hang at AB(28)-AD(30)
BQ_START_COL    = 20   # T
CRASH_START_COL = 24   # X
HANG_START_COL  = 28   # AB


# ── Data fetching ──────────────────────────────────────────────────────────────

def fetch_bigquery(bq_start=None, bq_end=None):
    if not BQ_AVAILABLE:
        print("WARNING: google-cloud-bigquery not installed, skipping BQ fetch.", file=sys.stderr)
        return []
    _bq_start = bq_start or BQ_START
    _bq_end   = bq_end   or BQ_END
    query = f"""
        SELECT partition_date as dt, appId, count(distinct clientId) as user_count
        FROM `fulfillment-dwh-production.curated_data_shared_coredata_tracking.perseus_events_rider_app`
        WHERE partition_date BETWEEN '{_bq_start}' AND '{_bq_end}'
          AND platform = 'iOS'
        GROUP BY ALL
        ORDER BY appId, dt ASC
    """
    try:
        client = bq_client.Client(project="logistics-rider-staging")
        rows = list(client.query(query))
        return [
            {
                "dt":         str(row.dt),
                "appId":      row.appId or "",
                "user_count": row.user_count,
            }
            for row in rows
        ]
    except Exception as exc:
        print(f"WARNING: BigQuery fetch failed, skipping BQ data: {exc}", file=sys.stderr)
        return []


def fetch_bigquery_by_version(version, bq_start=None, bq_end=None):
    """Return (raw_rows, brand_totals) for a specific appVersionCode.

    Runs one query per brand with LOWER(appId) LIKE and appVersionCode filters —
    matching the Dashboard SUMIFS approach. raw_rows is [{dt, appId, user_count}]
    suitable for writing to a Dashboard-style raw data section. brand_totals is
    {bq_key_lower: total_user_count} for AQS computation.
    """
    if not BQ_AVAILABLE:
        return [], {}
    _bq_start = bq_start or BQ_START
    _bq_end   = bq_end   or BQ_END
    all_rows     = []
    brand_totals = {}
    for brand in BRANDS:
        bkey = brand["bq_key"].lower()
        query = f"""
            SELECT partition_date as dt, appId, count(distinct clientId) as user_count
            FROM `fulfillment-dwh-production.curated_data_shared_coredata_tracking.perseus_events_rider_app`
            WHERE partition_date BETWEEN '{_bq_start}' AND '{_bq_end}'
              AND platform = 'iOS'
              AND appVersionCode = '{version}'
              AND LOWER(appId) LIKE '%{bkey}%'
            GROUP BY ALL
            ORDER BY appId, dt ASC
        """
        try:
            client = bq_client.Client(project="logistics-rider-staging")
            rows = list(client.query(query))
            brand_rows = [{"dt": str(r.dt), "appId": r.appId or "", "user_count": int(r.user_count or 0)} for r in rows]
            total = sum(r["user_count"] for r in brand_rows)
        except Exception as exc:
            print(f"WARNING: BQ version={version} brand={bkey}: {exc}", file=sys.stderr)
            brand_rows = []
            total = 0
        all_rows.extend(brand_rows)
        brand_totals[bkey] = total
        print(f"    {bkey}: {total:,} users")
    return all_rows, brand_totals


def fetch_firebase_frames(bq_start=None, bq_end=None):
    """Return {"frozen": float, "skipped": float} aggregated across ALL brands.

    Runs a single query that UNIONs all brand tables — same structure as the
    reference SQL — and returns one row.  The same values are applied to every
    brand row in the Consolidation sheet.
    """
    if not BQ_AVAILABLE:
        print("WARNING: google-cloud-bigquery not available, skipping Firebase frames fetch.",
              file=sys.stderr)
        return {}

    _bq_start = bq_start or BQ_START
    _bq_end   = bq_end   or BQ_END

    selects = "\n  UNION ALL".join(
        f"""
  SELECT
    app_display_version, app_build_version, os_version, device_name, country,
    event_name, event_type, event_timestamp,
    trace_info.duration_us,
    trace_info.screen_info,
    trace_info.metric_info
  FROM `logistics-54934.firebase_performance.{table}`
  WHERE TIMESTAMP_TRUNC(_PARTITIONTIME, DAY) >= TIMESTAMP('{_bq_start}')
    AND TIMESTAMP_TRUNC(_PARTITIONTIME, DAY) <= TIMESTAMP('{_bq_end}')"""
        for _, table in FIREBASE_BRANDS
    )

    query = f"""
WITH fb_performance AS ({selects}
),
frames_grader AS (
  SELECT
    event_name,
    event_type,
    duration_us,
    screen_info.slow_frame_ratio,
    screen_info.frozen_frame_ratio,
    CONCAT(event_name, CAST(event_timestamp AS STRING), app_build_version, country) AS event_id,
    CASE
      WHEN COUNT(CONCAT(event_name, CAST(event_timestamp AS STRING), app_build_version, country)) > 0
        THEN 1 ELSE 0
    END AS total_event_count,
    CASE WHEN COUNTIF(screen_info.slow_frame_ratio   > 0.5)   > 0 THEN 1 ELSE 0 END AS count_skipped_greater50,
    CASE WHEN COUNTIF(screen_info.frozen_frame_ratio > 0.001) > 0 THEN 1 ELSE 0 END AS count_frozen_greater1
  FROM fb_performance
  WHERE screen_info IS NOT NULL
    AND CHAR_LENGTH(event_name) > 9
  GROUP BY 1, 2, 3, 4, 5, 6
),
events_totals AS (
  SELECT
    SUM(total_event_count)       AS events_count,
    SUM(count_skipped_greater50) AS skipped_frames_count,
    SUM(count_frozen_greater1)   AS frozen_frames_count
  FROM frames_grader
),
frames_percentages_calc AS (
  SELECT
    (skipped_frames_count / events_count) * 100 AS skipped_frames_percentage,
    (frozen_frames_count  / events_count) * 100 AS forzen_frames_percentage
  FROM events_totals
)
SELECT
  ROUND(skipped_frames_percentage, 2) AS skipped_frames_percentage,
  ROUND(forzen_frames_percentage,  2) AS forzen_frames_percentage,
  ROUND(100 - ((0.2 * skipped_frames_percentage) + (0.8 * forzen_frames_percentage)), 2) AS performance_score
FROM frames_percentages_calc
"""
    try:
        client = bq_client.Client(project="logistics-rider-staging")
        rows = list(client.query(query))
        if not rows:
            print("WARNING: Firebase frames query returned no rows.", file=sys.stderr)
            return {}
        row = rows[0]
        result = {
            "frozen":  float(row.forzen_frames_percentage),
            "skipped": float(row.skipped_frames_percentage),
        }
        print(f"Firebase frames: frozen={result['frozen']}%  skipped={result['skipped']}%")
        return result
    except Exception as exc:
        print(f"ERROR fetching Firebase frames: {exc}", file=sys.stderr)
        return {}


def fetch_firebase_frames_by_version(version, bq_start=None, bq_end=None):
    """Return {"frozen": float, "skipped": float, "score": float} for a specific app_display_version.

    Runs the same UNION query as fetch_firebase_frames but adds an app_display_version
    filter to each brand table, isolating frame metrics for a single release.
    """
    if not BQ_AVAILABLE:
        print(f"WARNING: BigQuery unavailable, skipping frames fetch for version {version}.",
              file=sys.stderr)
        return {}

    _bq_start = bq_start or BQ_START
    _bq_end   = bq_end   or BQ_END

    selects = "\n  UNION ALL".join(
        f"""
  SELECT
    app_display_version, app_build_version, os_version, device_name, country,
    event_name, event_type, event_timestamp,
    trace_info.duration_us,
    trace_info.screen_info,
    trace_info.metric_info
  FROM `logistics-54934.firebase_performance.{table}`
  WHERE TIMESTAMP_TRUNC(_PARTITIONTIME, DAY) >= TIMESTAMP('{_bq_start}')
    AND TIMESTAMP_TRUNC(_PARTITIONTIME, DAY) <= TIMESTAMP('{_bq_end}')
    AND app_display_version LIKE '{version}%'"""
        for _, table in FIREBASE_BRANDS
    )

    query = f"""
WITH fb_performance AS ({selects}
),
frames_grader AS (
  SELECT
    event_name,
    event_type,
    duration_us,
    screen_info.slow_frame_ratio,
    screen_info.frozen_frame_ratio,
    CONCAT(event_name, CAST(event_timestamp AS STRING), app_build_version, country) AS event_id,
    CASE
      WHEN COUNT(CONCAT(event_name, CAST(event_timestamp AS STRING), app_build_version, country)) > 0
        THEN 1 ELSE 0
    END AS total_event_count,
    CASE WHEN COUNTIF(screen_info.slow_frame_ratio   > 0.5)   > 0 THEN 1 ELSE 0 END AS count_skipped_greater50,
    CASE WHEN COUNTIF(screen_info.frozen_frame_ratio > 0.001) > 0 THEN 1 ELSE 0 END AS count_frozen_greater1
  FROM fb_performance
  WHERE screen_info IS NOT NULL
    AND CHAR_LENGTH(event_name) > 9
  GROUP BY 1, 2, 3, 4, 5, 6
),
events_totals AS (
  SELECT
    SUM(total_event_count)       AS events_count,
    SUM(count_skipped_greater50) AS skipped_frames_count,
    SUM(count_frozen_greater1)   AS frozen_frames_count
  FROM frames_grader
),
frames_percentages_calc AS (
  SELECT
    (skipped_frames_count / events_count) * 100 AS skipped_frames_percentage,
    (frozen_frames_count  / events_count) * 100 AS forzen_frames_percentage
  FROM events_totals
)
SELECT
  ROUND(skipped_frames_percentage, 2) AS skipped_frames_percentage,
  ROUND(forzen_frames_percentage,  2) AS forzen_frames_percentage,
  ROUND(100 - ((0.2 * skipped_frames_percentage) + (0.8 * forzen_frames_percentage)), 2) AS performance_score
FROM frames_percentages_calc
"""
    try:
        client = bq_client.Client(project="logistics-rider-staging")
        rows = list(client.query(query))
        if not rows:
            print(f"WARNING: No Firebase frame data for version {version}.", file=sys.stderr)
            return {}
        row = rows[0]
        frozen_val  = row.forzen_frames_percentage
        skipped_val = row.skipped_frames_percentage
        score_val   = row.performance_score
        if frozen_val is None or skipped_val is None:
            print(f"WARNING: Firebase frames for v{version} returned NULL (no events in period).",
                  file=sys.stderr)
            return {}
        result = {
            "frozen":  float(frozen_val),
            "skipped": float(skipped_val),
            "score":   float(score_val) if score_val is not None else None,
        }
        print(f"  v{version}: frozen={result['frozen']}%  skipped={result['skipped']}%  score={result['score']}")
        return result
    except Exception as exc:
        print(f"ERROR fetching Firebase frames for version {version}: {exc}", file=sys.stderr)
        return {}


def fetch_sentry_users_per_version(rel, base_query, environment=None):
    """Return count_unique(user) from Sentry for a specific release/dist.

    Appends the dist/release filter from rel_filter_str(rel) to base_query and
    calls the Discover /events/ endpoint for a single aggregated user count.
    """
    dist_filter = rel_filter_str(rel)
    query = f"{base_query} {dist_filter}".strip()

    params = [
        ("project", PROJECT),
        ("query",   query),
        ("start",   START),
        ("end",     END),
        ("field",   "count_unique(user)"),
        ("dataset", "errors"),
    ]
    if environment:
        params.append(("environment", environment))

    url = (f"https://sentry.io/api/0/organizations/{ORG}/events/?"
           + urllib.parse.urlencode(params))
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read())
        rows = data.get("data", [])
        if not rows:
            return 0
        return int(rows[0].get("count_unique(user)", 0) or 0)
    except Exception as exc:
        print(f"ERROR fetching Sentry users for version {rel['version']}: {exc}", file=sys.stderr)
        return 0


def fetch_sentry_users_per_version_per_brand(rel, base_query, environment=None):
    """Return {sentry_key_lower: count_unique(user)} per brand for a specific release/dist."""
    dist_filter = rel_filter_str(rel)
    result = {}
    for brand in BRANDS:
        skey = brand["sentry_key"]
        brand_query = f"{base_query} {dist_filter} ((Brand:\"\" brand:*{skey}*) OR Brand:*{skey}*)".strip()
        params = [
            ("project", PROJECT),
            ("query",   brand_query),
            ("start",   START),
            ("end",     END),
            ("field",   "count_unique(user)"),
            ("dataset", "errors"),
        ]
        if environment:
            params.append(("environment", environment))
        url = (f"https://sentry.io/api/0/organizations/{ORG}/events/?"
               + urllib.parse.urlencode(params))
        req = urllib.request.Request(url, headers=HEADERS)
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read())
            rows = data.get("data", [])
            count = int(rows[0].get("count_unique(user)", 0) or 0) if rows else 0
        except Exception as exc:
            print(f"WARNING: brand={skey} v{rel['version']}: {exc}", file=sys.stderr)
            count = 0
        result[skey.lower()] = count
        time.sleep(0.3)
    return result


def fetch_discover(query, environment=None, start=None, end=None, brand_field="Brand"):
    rows = []
    cursor = None
    _start = start or START
    _end   = end   or END

    while True:
        params = [
            ("project",  PROJECT),
            ("query",    query),
            ("start",    _start),
            ("end",      _end),
            ("field",    "timestamp.to_day"),
            ("field",    "brand"),
            ("field",    "Brand"),
            ("field",    "count_unique(user)"),
            ("sort",     "timestamp.to_day"),
            ("dataset",  "errors"),
            ("per_page", "50"),
        ]
        if environment:
            params.append(("environment", environment))
        if cursor:
            params.append(("cursor", cursor))

        url = (f"https://sentry.io/api/0/organizations/{ORG}/events/?"
               + urllib.parse.urlencode(params))
        req = urllib.request.Request(url, headers=HEADERS)
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read())
                link = resp.headers.get("Link", "")
        except Exception as e:
            print(f"ERROR fetching Discover: {e}", file=sys.stderr)
            break

        rows.extend(data.get("data", []))

        next_cursor = None
        for part in link.split(","):
            if 'rel="next"' in part and 'results="true"' in part:
                for seg in part.split(";"):
                    seg = seg.strip()
                    if seg.startswith("<"):
                        next_cursor = urllib.parse.parse_qs(
                            urllib.parse.urlparse(seg[1:-1]).query
                        ).get("cursor", [None])[0]
        if not next_cursor:
            break
        cursor = next_cursor

    return rows


def fetch_discover_max(query, environment=None, start=None, end=None):
    """Run the Discover query twice and return rows with the max count_unique(user)
    per (Brand, day) pair, so intermittent Sentry under-counts are corrected."""
    run1 = fetch_discover(query, environment=environment, start=start, end=end)
    run2 = fetch_discover(query, environment=environment, start=start, end=end)

    best = {}  # (brand, day) → (max_count, row)
    for rows in (run1, run2):
        for row in rows:
            brand = (row.get("Brand") or "")
            day   = (row.get("timestamp.to_day") or "")[:10]
            count = int(row.get("count_unique(user)", 0) or 0)
            key   = (brand.lower(), day)
            if key not in best or count > best[key][0]:
                best[key] = (count, row)

    result = []
    for (_, _), (count, row) in best.items():
        merged = dict(row)
        merged["count_unique(user)"] = count
        result.append(merged)
    return result


def fetch_discover_per_brand(query, environment=None, start=None, end=None):
    """Run a separate Discover query per brand with an explicit Brand filter.
    Merges all results into a single list. More reliable than a single query
    since Sentry can silently drop or zero-out brand rows in grouped results."""
    all_rows = []
    for brand in BRANDS:
        skey = brand["sentry_key"]
        brand_query = f'{query} ((Brand:"" brand:*{skey}*) OR Brand:*{skey}*)'
        rows = fetch_discover(brand_query, environment=environment, start=start, end=end)
        total = sum(int(r.get("count_unique(user)", 0) or 0) for r in rows)
        print(f"    {skey}: {total} users across {len(rows)} day-rows")
        all_rows.extend(rows)
    return all_rows


def fetch_discover_per_brand_for_release(rel, base_query, environment=None, start=None, end=None):
    """Like fetch_discover_per_brand but adds the release dist filter to each brand query."""
    dist_filter = rel_filter_str(rel)
    excluded = {b.lower() for b in rel.get("excluded_brands", [])}
    all_rows = []
    for brand in BRANDS:
        skey = brand["sentry_key"]
        if skey.lower() in excluded:
            print(f"    {skey}: skipped (not rolled out for v{rel['version']})")
            continue
        brand_query = f'{base_query} {dist_filter} ((Brand:"" brand:*{skey}*) OR Brand:*{skey}*)'
        rows = fetch_discover(brand_query, environment=environment, start=start, end=end)
        # Tag each row with the brand we queried for — the brand field in the Sentry response
        # may be empty or a short name without the full package path, so we carry the key
        # explicitly to guarantee correct aggregation later.
        for r in rows:
            r["_sentry_key"] = skey.lower()
        total = sum(int(r.get("count_unique(user)", 0) or 0) for r in rows)
        print(f"    {skey}: {total} users across {len(rows)} day-rows")
        all_rows.extend(rows)
        time.sleep(0.3)
    return all_rows


def shape_rows(raw, user_col):
    rows = [
        {
            user_col:           int(row.get("count_unique(user)", 0)),
            "day":              row.get("timestamp.to_day", "")[:10],
            "timestamp.to_day": row.get("timestamp.to_day", ""),
            "Brand":            row.get("Brand") or row.get("brand") or "",
            "_sentry_key":      row.get("_sentry_key", ""),
        }
        for row in raw
    ]
    return sorted(rows, key=lambda r: r["timestamp.to_day"])


def aggregate_by_brand(rows, user_col="count_unique(user)"):
    """Aggregate user counts from Sentry rows by brand.

    Accepts either raw Discover rows (user_col="count_unique(user)") or
    shaped rows (user_col="CRASH_USERS" / "HANG_USERS").
    Returns dict: {sentry_key_lower: total_count} for each brand in BRANDS.

    Prefers the _sentry_key tag (set by fetch_discover_per_brand_for_release) over
    substring-matching the Brand field, so results are correct even when Sentry returns
    a short name without the full package path.
    """
    totals = {}
    for brand in BRANDS:
        skey = brand["sentry_key"].lower()
        total = 0
        for row in rows:
            tagged_key = row.get("_sentry_key", "")
            if tagged_key:
                matches = (tagged_key == skey)
            else:
                brand_val = (row.get("Brand") or "").lower()
                matches = bool(brand_val) and skey in brand_val
            if matches:
                total += int(row.get(user_col, 0) or 0)
        totals[skey] = total
    return totals


def backfill_zero_rows(rows, user_col, query, environment=None):
    """For any row where user_col is 0, re-query Sentry for that day to get the real count."""
    zero_days = {row["day"] for row in rows if row[user_col] == 0}
    for day in sorted(zero_days):
        print(f"  Re-querying {user_col} for {day} (row had 0 count)...")
        day_raw = fetch_discover(
            query, environment=environment,
            start=f"{day}T00:00:00.000",
            end=f"{day}T23:59:59.999",
        )
        day_counts = {}
        for r in day_raw:
            brand = (r.get("Brand") or r.get("brand") or "").lower()
            day_counts[brand] = day_counts.get(brand, 0) + int(r.get("count_unique(user)", 0) or 0)
        for row in rows:
            if row["day"] == day and row[user_col] == 0:
                brand = (row.get("Brand") or row.get("brand") or "").lower()
                row[user_col] = day_counts.get(brand, 0)


# ── AQS formula ────────────────────────────────────────────────────────────────

def aqs_score(value, col_key):
    """Compute AQS score for a single metric value."""
    cfg = AQS_CONFIG[col_key]
    baseline = cfg["baseline"]
    target   = cfg["target"]
    weight   = cfg["weight"]
    score = min(100, max(0, (((value - baseline) / (target - baseline)) * 50) + 50))
    return score * weight / 100


# ── Color coding (for HTML report) ────────────────────────────────────────────

def color_cfu(v):
    if v >= 99.9:
        return "#E8F5EA"
    elif v >= 99.7:
        return "#FFF8E1"
    return "#FFE4E6"


def color_hang(v):
    return color_cfu(v)  # same thresholds


def color_app_size(v):
    if v <= 60:
        return "#E8F5EA"
    elif v <= 75:
        return "#FFF8E1"
    return "#FFE4E6"


def color_asti(v):
    if v <= 2:
        return "#E8F5EA"
    elif v <= 4:
        return "#FFF8E1"
    return "#FFE4E6"


def color_stti(v):
    if v <= 0.5:
        return "#E8F5EA"
    elif v <= 1.5:
        return "#FFF8E1"
    return "#FFE4E6"


def color_frozen(v):
    if v <= 1:
        return "#E8F5EA"
    elif v <= 3:
        return "#FFF8E1"
    return "#FFE4E6"


def color_skipped(v):
    if v <= 1:
        return "#E8F5EA"
    elif v <= 2:
        return "#FFF8E1"
    return "#FFE4E6"


COLOR_FNS = {
    "cfu":      color_cfu,
    "hang":     color_hang,
    "app_size": color_app_size,
    "asti":     color_asti,
    "stti":     color_stti,
    "frozen":   color_frozen,
    "skipped":  color_skipped,
}

# Pastel brand colors — soft, muted tones
BRAND_COLORS = [
    "#FFE8EA",  # Foodpanda    — Soft Rose
    "#E6FAF0",  # Foodora      — Soft Mint
    "#E4F1FF",  # Talabat      — Soft Sky Blue
    "#FEFEE8",  # pedidosya    — Soft Lemon
    "#F2E8F6",  # HungerStation — Soft Lavender
    "#E4F9F8",  # Yemeksepeti  — Soft Turquoise
    "#FEFEF2",  # Glovo        — Soft Cream
    "#DFF8F7",  # Woowa        — Soft Aqua
    "#EEE8E2",  # efood        — Soft Taupe
    "#EAE4FA",  # Foody        — Soft Purple
]


# ── HTML report generation ─────────────────────────────────────────────────────

def _build_version_aqs_panel_html(version_aqs_data, date_range):
    """Build the HTML content for the Version AQS panel.

    Each entry in version_aqs_data must contain:
        version, cfu, hang, asti, stti, app_size, frozen, skipped,
        aqs_scores (dict), final_aqs (float), crash_users, hang_users,
        no_data (bool)
    """
    if not version_aqs_data:
        return """
<div style="padding:40px;text-align:center;color:#9ca3af;font-size:13px;">
  No version AQS data available. Ensure Sentry and BigQuery are accessible when running the report.
</div>"""

    AQS_COL_ORDER = ["cfu", "hang", "app_size", "asti", "stti", "frozen", "skipped"]

    def aqs_bg(score):
        if score is None:
            return "#f9fafb"
        if score >= 80:
            return "#E8F5EA"
        if score >= 60:
            return "#FFF8E1"
        return "#FFE4E6"

    brand_rows_html = ""
    aqs_score_cells = ""
    version_colors = [
        "#E4F1FF", "#E6FAF0", "#FEFEE8", "#F2E8F6", "#E4F9F8",
    ]

    for i, vd in enumerate(version_aqs_data):
        ver     = vd.get("version", "—")
        no_data = vd.get("no_data", False)
        bg      = version_colors[i % len(version_colors)]
        cs      = f"background:{bg};"

        if no_data:
            brand_rows_html += f"""
      <tr style="background:{bg};">
        <td style="{cs}font-weight:600;">{ver}</td>
        <td colspan="8" style="{cs}text-align:center;color:#9ca3af;">No data available for this period</td>
      </tr>"""
        else:
            cfu      = vd.get("cfu",      0.0)
            hang     = vd.get("hang",     0.0)
            asti     = vd.get("asti",     0.0)
            stti     = vd.get("stti",     0.0)
            app_size = vd.get("app_size", 0.0)
            frozen   = vd.get("frozen",   0.0)
            skipped  = vd.get("skipped",  0.0)
            final    = vd.get("final_aqs", 0.0)
            brand_rows_html += f"""
      <tr style="background:{bg};">
        <td style="{cs}font-weight:600;white-space:nowrap;">{ver}</td>
        <td style="{cs}text-align:right;">{cfu:.2f}%</td>
        <td style="{cs}text-align:right;">{hang:.2f}%</td>
        <td style="{cs}text-align:right;">{app_size}</td>
        <td style="{cs}text-align:right;">{asti}</td>
        <td style="{cs}text-align:right;">{stti}</td>
        <td style="{cs}text-align:right;">{frozen:.2f}%</td>
        <td style="{cs}text-align:right;">{skipped:.2f}%</td>
        <td style="{cs}text-align:right;font-weight:700;">{final}</td>
      </tr>"""

    # AQS Score component row (sum of all versions' components averaged, or show per-version)
    # Show one row per version for AQS components — use a compact sub-table style
    aqs_rows_html = ""
    for vd in version_aqs_data:
        if vd.get("no_data"):
            continue
        scores = vd.get("aqs_scores", {})
        cells  = "".join(
            f'<td style="text-align:right;">{scores.get(k, 0):.4f}</td>'
            for k in AQS_COL_ORDER
        )
        aqs_rows_html += f"""
      <tr class="row-aqs">
        <td style="text-align:left;">{vd['version']}</td>
        {cells}
        <td style="text-align:right;font-weight:700;">{vd.get('final_aqs', 0)}</td>
      </tr>"""

    # Score cards
    score_cards_html = '<div class="score-wrap">'
    for vd in version_aqs_data:
        if vd.get("no_data"):
            continue
        score_cards_html += f"""
  <div class="score-card">
    <div class="sc-label">v{vd['version']}</div>
    <div class="sc-value">{vd.get('final_aqs', 0)}</div>
    <div class="sc-sub">AQS Score</div>
  </div>"""
    score_cards_html += "\n</div>"

    return f"""
<p style="font-size:13px;color:#6b7280;margin-bottom:20px;">
  Full AQS score per pinned release version. Crashes &amp; hangs from Sentry (dist filter);
  frozen &amp; skipped frames from Firebase Performance BigQuery (version filter);
  ASTI, STTI, and app size use current fleet values.
  Period: <strong>{date_range}</strong>
</p>

<div class="table-wrap">
  <table>
    <thead>
      <tr>
        <th>Version</th>
        <th class="num">Crash Free %</th>
        <th class="num">Hang Free %</th>
        <th class="num">App Size (MB)</th>
        <th class="num">ASTI (s)</th>
        <th class="num">STTI (s)</th>
        <th class="num">Frozen Frames %</th>
        <th class="num">Skipped Frames %</th>
        <th class="num">Final AQS</th>
      </tr>
    </thead>
    <tbody>
      {brand_rows_html}
    </tbody>
  </table>
</div>

<div style="margin-bottom:20px;">
  <p style="font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.05em;color:#374151;margin-bottom:10px;">AQS Score Components (35% / 25% / 2% / 10% / 10% / 13% / 5%)</p>
  <div class="table-wrap">
    <table>
      <thead>
        <tr>
          <th>Version</th>
          <th class="num">Crash Free</th>
          <th class="num">Hang Free</th>
          <th class="num">App Size</th>
          <th class="num">ASTI</th>
          <th class="num">STTI</th>
          <th class="num">Frozen</th>
          <th class="num">Skipped</th>
          <th class="num">Final AQS</th>
        </tr>
      </thead>
      <tbody>
        {aqs_rows_html}
      </tbody>
    </table>
  </div>
</div>

{score_cards_html}

<div class="footer">
  &bull; <strong>Crash Free % / Hang Free %</strong> — Sentry <code>count_unique(user)</code>
    with dist filter, divided by total BQ iOS user-days for the period.<br>
  &bull; <strong>Frozen Frames % / Skipped Frames %</strong> — Firebase Performance BigQuery
    filtered by <code>app_display_version</code>.<br>
  &bull; <strong>ASTI, STTI, App Size</strong> — rider-weighted fleet averages (static per report run).<br>
  &bull; <strong>AQS formula</strong> — <code>min(100, max(0, (((value − baseline) / (target − baseline)) × 50) + 50)) × weight%</code><br>
  &bull; Query period: {date_range}
</div>"""


def _compute_metrics(bq_users_by_brand, crash_by_brand, hang_by_brand, firebase_data, start_date, end_date):
    """Returns (brand_metrics, avg_metrics, aqs_scores, final_aqs)."""
    days_in_period = (end_date - start_date).days + 1
    brand_metrics = []
    for brand in BRANDS:
        skey = brand["sentry_key"].lower()
        bkey = brand["bq_key"].lower()
        users = 0
        for appid, count in bq_users_by_brand.items():
            if bkey in appid:
                users += count
        crashes = crash_by_brand.get(skey, 0)
        hangs   = hang_by_brand.get(skey, 0)
        if users > 0:
            cfu  = max(0.0, min(100.0, int((1 - crashes / users) * 10000) / 100))
            hang = max(0.0, min(100.0, int((1 - hangs / users) * 10000) / 100))
        else:
            fallback = brand["riders"] * days_in_period
            cfu  = max(0.0, min(100.0, int((1 - crashes / fallback) * 10000) / 100))
            hang = max(0.0, min(100.0, int((1 - hangs / fallback) * 10000) / 100))
        brand_metrics.append({
            "name":     brand["name"],
            "riders":   brand["riders"],
            "users":    users,
            "crashes":  crashes,
            "hangs":    hangs,
            "cfu":      cfu,
            "hang":     hang,
            "app_size": brand["app_size"],
            "asti":     brand["asti"],
            "stti":     brand["stti"],
            "frozen":   firebase_data.get("frozen",  0.58),
            "skipped":  firebase_data.get("skipped", 1.2),
            "fallback": users == 0,
        })

    def weighted_avg(key):
        return int(sum(m[key] * w for m, w in zip(brand_metrics, WEIGHTS)) * 100) / 100

    avg_metrics = {k: weighted_avg(k) for k in ["cfu", "hang", "app_size", "asti", "stti", "frozen", "skipped"]}
    aqs_scores  = {k: round(aqs_score(avg_metrics[k], k), 4) for k in AQS_CONFIG}
    final_aqs   = round(sum(aqs_scores.values()), 2)
    return brand_metrics, avg_metrics, aqs_scores, final_aqs


def generate_html_report(bq_users_by_brand, crash_by_brand, hang_by_brand, firebase_data,
                          prev_bq_users=None, prev_crash=None, prev_hang=None, prev_firebase=None,
                          version_aqs_data=None):
    """Generate consolidation_report/index.html from pre-fetched data.

    Args:
        bq_users_by_brand: dict {appid_lower: total_user_count}
        crash_by_brand:    dict {sentry_key_lower: total_crash_users}
        hang_by_brand:     dict {sentry_key_lower: total_hang_users}
        firebase_data:     dict {"frozen": float, "skipped": float} (may be empty)
        prev_bq_users:     same shape as bq_users_by_brand for previous month (optional)
        prev_crash:        same shape as crash_by_brand for previous month (optional)
        prev_hang:         same shape as hang_by_brand for previous month (optional)
        prev_firebase:     same shape as firebase_data for previous month (optional)
        version_aqs_data:  list of per-version AQS dicts — {version, cfu, hang, asti, stti,
                           app_size, frozen, skipped, aqs_scores, final_aqs, ...} (optional)
    """
    OUT_DIR = "consolidation_report"

    fetch_date = TODAY.strftime("%-d %b %Y")
    date_range = f"{START_DATE.strftime('%-d %b')} – {END_DATE.strftime('%-d %b %Y')}"

    # ── Two-month path ─────────────────────────────────────────────────────────
    if prev_bq_users is not None:
        curr_metrics, curr_avg, curr_aqs, curr_final = _compute_metrics(
            bq_users_by_brand, crash_by_brand, hang_by_brand, firebase_data, START_DATE, END_DATE
        )
        prev_metrics, prev_avg, prev_aqs, prev_final = _compute_metrics(
            prev_bq_users, prev_crash, prev_hang, prev_firebase or {}, PREV_START_DATE, PREV_END_DATE
        )

        print("\nBrand metrics:")
        for m in curr_metrics:
            flag = " (fallback)" if m["fallback"] else ""
            print(f"  {m['name']:15s}  users={m['users']:>8,}  crashes={m['crashes']:>5}  hangs={m['hangs']:>5}"
                  f"  cfu={m['cfu']:.2f}%  hang={m['hang']:.2f}%{flag}")

        print(f"\nWeighted averages: {curr_avg}")
        print(f"AQS scores:        {curr_aqs}")
        print(f"FINAL AQS:         {curr_final}")

        print(f"\nPrev month brand metrics ({PREV_START_DATE.strftime('%-d %b')} – {PREV_END_DATE.strftime('%-d %b %Y')}):")
        for m in prev_metrics:
            flag = " (fallback)" if m["fallback"] else ""
            print(f"  {m['name']:15s}  users={m['users']:>8,}  crashes={m['crashes']:>5}  hangs={m['hangs']:>5}"
                  f"  cfu={m['cfu']:.2f}%  hang={m['hang']:.2f}%{flag}")

        print(f"\nPrev weighted averages: {prev_avg}")
        print(f"Prev AQS scores:        {prev_aqs}")
        print(f"PREV FINAL AQS:         {prev_final}")

        prev_month_label = PREV_START_DATE.strftime("%B %Y")
        curr_month_label = START_DATE.strftime("%B %Y")
        prev_date_range  = f"{PREV_START_DATE.strftime('%-d %b')} – {PREV_END_DATE.strftime('%-d %b %Y')}"
        curr_date_range  = date_range

        def _build_aqs_table_html(brand_metrics, avg_metrics, aqs_scores, final_aqs, period_date_range):
            fmt_pct = lambda v: f"{v:.2f}%"
            fmt_mb  = lambda v: f"{v}"
            fmt_s   = lambda v: f"{v}"

            brand_rows_html = ""
            for i, m in enumerate(brand_metrics):
                row_color = BRAND_COLORS[i % len(BRAND_COLORS)]
                fallback_note = (' <span style="font-size:10px;color:#6b7280;" title="BQ unavailable — using fallback">*</span>'
                                 if m["fallback"] else "")
                cell_style = f'background:{row_color};'
                brand_rows_html += f"""
      <tr style="background:{row_color};">
        <td style="{cell_style}font-weight:600;white-space:nowrap;">{m['name']}{fallback_note}</td>
        <td style="{cell_style}text-align:right;">{fmt_pct(m['cfu'])}</td>
        <td style="{cell_style}text-align:right;">{fmt_pct(m['hang'])}</td>
        <td style="{cell_style}text-align:right;">{fmt_mb(m['app_size'])}</td>
        <td style="{cell_style}text-align:right;">{fmt_s(m['asti'])}</td>
        <td style="{cell_style}text-align:right;">{fmt_s(m['stti'])}</td>
        <td style="{cell_style}text-align:right;">{fmt_pct(m['frozen'])}</td>
        <td style="{cell_style}text-align:right;">{fmt_pct(m['skipped'])}</td>
        <td style="{cell_style}text-align:right;">{m['riders']:,}</td>
      </tr>"""

            aqs_col_order = ["cfu", "hang", "app_size", "asti", "stti", "frozen", "skipped"]
            aqs_cells_html = "".join(
                f'<td style="text-align:right;">{aqs_scores[k]:.4f}</td>'
                for k in aqs_col_order
            )

            return f"""
<div class="table-wrap">
  <table>
    <thead>
      <tr>
        <th>Brand</th>
        <th class="num">Crash Free %</th>
        <th class="num">App Hangs %</th>
        <th class="num">App Size (MB)</th>
        <th class="num">ASTI (s)</th>
        <th class="num">STTI (s)</th>
        <th class="num">Frozen Frames %</th>
        <th class="num">Skipped Frames %</th>
        <th class="num">Riders</th>
      </tr>
    </thead>
    <tbody>
      {brand_rows_html}
      <tr class="row-avg">
        <td>Weighted AVG</td>
        <td>{avg_metrics['cfu']:.2f}%</td>
        <td>{avg_metrics['hang']:.2f}%</td>
        <td>{avg_metrics['app_size']}</td>
        <td>{avg_metrics['asti']}</td>
        <td>{avg_metrics['stti']}</td>
        <td>{avg_metrics['frozen']:.2f}%</td>
        <td>{avg_metrics['skipped']:.2f}%</td>
        <td></td>
      </tr>
      <tr class="row-aqs">
        <td>AQS Score</td>
        {aqs_cells_html}
        <td></td>
      </tr>
    </tbody>
  </table>
</div>

<div class="score-wrap">
  <div class="score-card">
    <div class="sc-label">Final AQS Score</div>
    <div class="sc-value">{final_aqs}</div>
    <div class="sc-sub">{period_date_range}</div>
  </div>
</div>"""

        prev_panel_html = _build_aqs_table_html(prev_metrics, prev_avg, prev_aqs, prev_final, prev_date_range)
        curr_panel_html = _build_aqs_table_html(curr_metrics, curr_avg, curr_aqs, curr_final, curr_date_range)
        version_panel_html = _build_version_aqs_panel_html(version_aqs_data or [], curr_date_range)

        bq_status_note = ""
        if not BQ_AVAILABLE:
            bq_status_note = """
    <div style="margin-top:12px;padding:10px 16px;background:#FFF3CD;border:1px solid #ffc107;
                border-radius:6px;font-size:13px;color:#856404;">
      ⚠ Using fallback data (BigQuery unavailable). Install <code>google-cloud-bigquery</code>
      and authenticate to fetch live data.
    </div>"""

        bq_available_str = 'True' if BQ_AVAILABLE else 'False'

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>iOS AQS Consolidation Report</title>
  <style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
            background: #f8fafc; color: #111; padding: 32px; }}
    h1 {{ font-size: 22px; font-weight: 700; margin-bottom: 4px; }}
    .subtitle {{ color: #666; font-size: 13px; margin-bottom: 24px; }}
    .warning-banner {{
      background: #fffbf0; border: 1px solid #f59e0b;
      border-left: 4px solid #f59e0b; border-radius: 6px;
      padding: 12px 16px; font-size: 13px; color: #92400e;
      font-weight: 600; margin-bottom: 16px;
    }}

    /* ── Tabs ── */
    .tabs {{ display: flex; gap: 4px; margin-bottom: 0; border-bottom: 2px solid #e5e7eb; }}
    .tab {{ padding: 9px 20px; font-size: 13px; font-weight: 600; cursor: pointer;
             border-radius: 6px 6px 0 0; border: 1px solid transparent;
             border-bottom: none; color: #6b7280; background: transparent;
             position: relative; top: 2px; transition: color .15s; }}
    .tab:hover {{ color: #111; }}
    .tab.active {{ color: #111; background: #fff; border-color: #e5e7eb;
                   border-bottom-color: #fff; }}
    .panel {{ display: none; background: #fff; border: 1px solid #e5e7eb;
               border-top: none; border-radius: 0 6px 6px 6px; padding: 28px; }}
    .panel.active {{ display: block; }}

    /* ── AQS Table panel ── */
    .score-wrap {{ display: flex; gap: 16px; margin-bottom: 28px; flex-wrap: wrap; }}
    .score-card {{ background: #f8fafc; border: 1px solid #e5e7eb; border-radius: 8px;
                   padding: 14px 24px; min-width: 180px; }}
    .score-card .sc-label {{ font-size: 11px; text-transform: uppercase;
                              letter-spacing: .05em; color: #6b7280; margin-bottom: 4px; }}
    .score-card .sc-value {{ font-size: 32px; font-weight: 700; color: #111; }}
    .score-card .sc-sub   {{ font-size: 11px; color: #9ca3af; margin-top: 2px; }}
    .table-wrap {{ overflow-x: auto; margin-bottom: 28px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    thead th {{
      text-align: left; padding: 8px 12px;
      border-bottom: 2px solid #e5e7eb; color: #666; font-weight: 600;
      font-size: 11px; text-transform: uppercase; letter-spacing: .05em; white-space: nowrap;
    }}
    thead th.num {{ text-align: right; }}
    tbody tr {{ border-bottom: 1px solid #f3f4f6; }}
    tbody tr:hover {{ background: #f9fafb; }}
    tbody td {{ padding: 9px 12px; vertical-align: middle; white-space: nowrap; text-align: right; }}
    tbody td:first-child {{ text-align: left; font-weight: 500; }}
    tr.row-avg td {{
      background: #f5fbff !important; font-weight: 700;
      border-top: 2px solid #dbeafe; color: #3b82f6;
    }}
    tr.row-avg td:first-child {{ text-align: left; }}
    tr.row-aqs td {{
      background: #f6fef8 !important; font-weight: 700;
      border-top: 2px solid #d1fae5; color: #34d399;
    }}
    tr.row-aqs td:first-child {{ text-align: left; }}

    .footer {{ margin-top: 32px; font-size: 11px; color: #9ca3af;
               border-top: 1px solid #e5e7eb; padding-top: 16px; line-height: 1.8; }}
    .footer code {{ background: #f3f4f6; padding: 1px 5px; border-radius: 3px; font-size: 11px; }}
  </style>
</head>
<body>

<h1>iOS AQS Consolidation Report</h1>
<p class="subtitle">
  Fetched: {fetch_date}
</p>

{bq_status_note}

<div class="tabs" id="tabs">
  <div class="tab active" data-panel="prev-month">{prev_month_label}</div>
  <div class="tab" data-panel="curr-month">{curr_month_label}</div>
  <div class="tab" data-panel="version-aqs">Version AQS</div>
</div>

<!-- Panel: prev month -->
<div class="panel active" id="panel-prev-month">
  {prev_panel_html}

  <div class="footer">
    <strong>Notes:</strong><br>
    &bull; <strong>Crash Free % / App Hangs %</strong> — computed from Sentry Discover API
      (<code>level:fatal handled:no !stack.package:*gpsmaster* !stack.package:*GPSTraveller* !user.id:*-*-*-*-* !issue:RIDER-APP-IOS-Z7 !issue:RIDER-APP-IOS-1BM !issue:RIDER-APP-IOS-3DVQ</code> / <code>!user.id:*-*-*-*-* app.in_foreground:True "*App hang* detected*"</code>)
      against BQ daily user counts for the period.<br>
    &bull; <strong>Frozen Frames % / Skipped Frames %</strong> — single aggregated value from Firebase Performance BQ,
      applied to all brands equally.<br>
    &bull; <strong>Weighted AVG</strong> — SUMPRODUCT of brand values with rider-share weights.<br>
    &bull; <strong>AQS formula</strong> — <code>min(100, max(0, (((value - baseline) / (target - baseline)) * 50) + 50)) * weight%</code><br>
    &bull; BQ_AVAILABLE = <code>{bq_available_str}</code> &nbsp;|&nbsp;
      Fallback frozen=0.58, skipped=1.2 used when Firebase BQ is unavailable.<br>
    &bull; Rows marked with <span style="color:#9ca3af">*</span> used fallback CFU/hang values (BQ user count was zero).
  </div>
</div>

<!-- Panel: curr month -->
<div class="panel" id="panel-curr-month">
  <div class="warning-banner">
    ⚠ App size, ASTI and STTI have the older values, not recently updated. Update these columns manually before sharing. For the accurate AQS score, update these values.
  </div>
  {curr_panel_html}

  <div class="footer">
    <strong>Notes:</strong><br>
    &bull; <strong>Crash Free % / App Hangs %</strong> — computed from Sentry Discover API
      (<code>level:fatal handled:no !stack.package:*gpsmaster* !stack.package:*GPSTraveller* !user.id:*-*-*-*-* !issue:RIDER-APP-IOS-Z7 !issue:RIDER-APP-IOS-1BM !issue:RIDER-APP-IOS-3DVQ</code> / <code>!user.id:*-*-*-*-* app.in_foreground:True "*App hang* detected*"</code>)
      against BQ daily user counts for the period.<br>
    &bull; <strong>Frozen Frames % / Skipped Frames %</strong> — single aggregated value from Firebase Performance BQ,
      applied to all brands equally.<br>
    &bull; <strong>App Size, ASTI, STTI</strong> — older values, not recently updated; update manually for accurate AQS score.<br>
    &bull; <strong>Weighted AVG</strong> — SUMPRODUCT of brand values with rider-share weights.<br>
    &bull; <strong>AQS formula</strong> — <code>min(100, max(0, (((value - baseline) / (target - baseline)) * 50) + 50)) * weight%</code><br>
    &bull; BQ_AVAILABLE = <code>{bq_available_str}</code> &nbsp;|&nbsp;
      Fallback frozen=0.58, skipped=1.2 used when Firebase BQ is unavailable.<br>
    &bull; Rows marked with <span style="color:#9ca3af">*</span> used fallback CFU/hang values (BQ user count was zero).
  </div>
</div>

<!-- Panel: version aqs -->
<div class="panel" id="panel-version-aqs">
  {version_panel_html}
</div>

<script>
document.getElementById('tabs').addEventListener('click', function(e) {{
  var t = e.target.closest('.tab');
  if (!t) return;
  document.querySelectorAll('.tab').forEach(function(x) {{ x.classList.remove('active'); }});
  document.querySelectorAll('.panel').forEach(function(x) {{ x.classList.remove('active'); }});
  t.classList.add('active');
  document.getElementById('panel-' + t.dataset.panel).classList.add('active');
}});
</script>

</body>
</html>"""

        os.makedirs(OUT_DIR, exist_ok=True)
        out_path = os.path.join(OUT_DIR, "index.html")
        with open(out_path, "w") as f:
            f.write(html)

        print(f"\nReport written to: {out_path}")
        return

    # ── Single-month path ──────────────────────────────────────────────────────
    brand_metrics, avg_metrics, aqs_scores, final_aqs = _compute_metrics(
        bq_users_by_brand, crash_by_brand, hang_by_brand, firebase_data, START_DATE, END_DATE
    )

    print("\nBrand metrics:")
    for m in brand_metrics:
        flag = " (fallback)" if m["fallback"] else ""
        print(f"  {m['name']:15s}  users={m['users']:>8,}  crashes={m['crashes']:>5}  hangs={m['hangs']:>5}"
              f"  cfu={m['cfu']:.2f}%  hang={m['hang']:.2f}%{flag}")

    print(f"\nWeighted averages: {avg_metrics}")
    print(f"AQS scores:        {aqs_scores}")
    print(f"FINAL AQS:         {final_aqs}")

    # ── HTML helpers ──────────────────────────────────────────────────────────
    def td(value, col_key, fmt=None, extra_style=""):
        """Return a <td> with background color coding and formatted value."""
        color = COLOR_FNS[col_key](value)
        text  = fmt(value) if fmt else str(value)
        style = f"background:{color};{extra_style}"
        return f'<td style="{style}">{text}</td>'

    def fmt_pct(v):
        return f"{v:.2f}%"

    def fmt_mb(v):
        return f"{v}"

    def fmt_s(v):
        return f"{v}"

    # ── Brand data rows — each brand gets its own pastel row color ────────────
    brand_rows_html = ""
    for i, m in enumerate(brand_metrics):
        row_color = BRAND_COLORS[i % len(BRAND_COLORS)]
        fallback_note = (' <span style="font-size:10px;color:#6b7280;" title="BQ unavailable — using fallback">*</span>'
                         if m["fallback"] else "")
        cell_style = f'background:{row_color};'
        brand_rows_html += f"""
      <tr style="background:{row_color};">
        <td style="{cell_style}font-weight:600;white-space:nowrap;">{m['name']}{fallback_note}</td>
        <td style="{cell_style}text-align:right;">{fmt_pct(m['cfu'])}</td>
        <td style="{cell_style}text-align:right;">{fmt_pct(m['hang'])}</td>
        <td style="{cell_style}text-align:right;">{fmt_mb(m['app_size'])}</td>
        <td style="{cell_style}text-align:right;">{fmt_s(m['asti'])}</td>
        <td style="{cell_style}text-align:right;">{fmt_s(m['stti'])}</td>
        <td style="{cell_style}text-align:right;">{fmt_pct(m['frozen'])}</td>
        <td style="{cell_style}text-align:right;">{fmt_pct(m['skipped'])}</td>
        <td style="{cell_style}text-align:right;">{m['riders']:,}</td>
      </tr>"""

    # ── AQS score row cells ────────────────────────────────────────────────────
    aqs_col_order = ["cfu", "hang", "app_size", "asti", "stti", "frozen", "skipped"]
    aqs_cells_html = "".join(
        f'<td style="text-align:right;">{aqs_scores[k]:.4f}</td>'
        for k in aqs_col_order
    )

    bq_status_note = ""
    if not BQ_AVAILABLE:
        bq_status_note = """
    <div style="margin-top:12px;padding:10px 16px;background:#FFF3CD;border:1px solid #ffc107;
                border-radius:6px;font-size:13px;color:#856404;">
      ⚠ Using fallback data (BigQuery unavailable). Install <code>google-cloud-bigquery</code>
      and authenticate to fetch live data.
    </div>"""

    brand_json = json.dumps([
        {
            "name":     m["name"],
            "color":    BRAND_COLORS[i],
            "cfu":      m["cfu"],
            "hang":     m["hang"],
            "app_size": m["app_size"],
            "asti":     m["asti"],
            "stti":     m["stti"],
            "frozen":   m["frozen"],
            "skipped":  m["skipped"],
            "riders":   m["riders"],
        }
        for i, m in enumerate(brand_metrics)
    ], indent=2)

    bq_available_str = 'True' if BQ_AVAILABLE else 'False'

    stale_warning_html = (
        '<div class="warning-banner">\n'
        '  ⚠ App size, ASTI and STTI have the older values, not recently updated.'
        ' Update these columns manually before sharing.'
        ' For the accurate AQS score, update these values.\n'
        '</div>'
        if TODAY.day >= 3 else ""
    )
    stale_footer_note = (
        '  &bull; <strong>App Size, ASTI, STTI</strong> — older values, not recently updated;'
        ' update manually for accurate AQS score.<br>\n'
        if TODAY.day >= 3 else ""
    )

    version_panel_html = _build_version_aqs_panel_html(version_aqs_data or [], date_range)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>iOS AQS Consolidation Report</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
  <style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
            background: #f8fafc; color: #111; padding: 32px; }}
    h1 {{ font-size: 22px; font-weight: 700; margin-bottom: 4px; }}
    .subtitle {{ color: #666; font-size: 13px; margin-bottom: 24px; }}
    .warning-banner {{
      background: #fffbf0; border: 1px solid #f59e0b;
      border-left: 4px solid #f59e0b; border-radius: 6px;
      padding: 12px 16px; font-size: 13px; color: #92400e;
      font-weight: 600; margin-bottom: 16px;
    }}

    /* ── Tabs ── */
    .tabs {{ display: flex; gap: 4px; margin-bottom: 0; border-bottom: 2px solid #e5e7eb; }}
    .tab {{ padding: 9px 20px; font-size: 13px; font-weight: 600; cursor: pointer;
             border-radius: 6px 6px 0 0; border: 1px solid transparent;
             border-bottom: none; color: #6b7280; background: transparent;
             position: relative; top: 2px; transition: color .15s; }}
    .tab:hover {{ color: #111; }}
    .tab.active {{ color: #111; background: #fff; border-color: #e5e7eb;
                   border-bottom-color: #fff; }}
    .panel {{ display: none; background: #fff; border: 1px solid #e5e7eb;
               border-top: none; border-radius: 0 6px 6px 6px; padding: 28px; }}
    .panel.active {{ display: block; }}

    /* ── AQS Table panel ── */
    .score-wrap {{ display: flex; gap: 16px; margin-bottom: 28px; flex-wrap: wrap; }}
    .score-card {{ background: #f8fafc; border: 1px solid #e5e7eb; border-radius: 8px;
                   padding: 14px 24px; min-width: 180px; }}
    .score-card .sc-label {{ font-size: 11px; text-transform: uppercase;
                              letter-spacing: .05em; color: #6b7280; margin-bottom: 4px; }}
    .score-card .sc-value {{ font-size: 32px; font-weight: 700; color: #111; }}
    .score-card .sc-sub   {{ font-size: 11px; color: #9ca3af; margin-top: 2px; }}
    .table-wrap {{ overflow-x: auto; margin-bottom: 28px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    thead th {{
      text-align: left; padding: 8px 12px;
      border-bottom: 2px solid #e5e7eb; color: #666; font-weight: 600;
      font-size: 11px; text-transform: uppercase; letter-spacing: .05em; white-space: nowrap;
    }}
    thead th.num {{ text-align: right; }}
    tbody tr {{ border-bottom: 1px solid #f3f4f6; }}
    tbody tr:hover {{ background: #f9fafb; }}
    tbody td {{ padding: 9px 12px; vertical-align: middle; white-space: nowrap; text-align: right; }}
    tbody td:first-child {{ text-align: left; font-weight: 500; }}
    tr.row-avg td {{
      background: #f5fbff !important; font-weight: 700;
      border-top: 2px solid #dbeafe; color: #3b82f6;
    }}
    tr.row-avg td:first-child {{ text-align: left; }}
    tr.row-aqs td {{
      background: #f6fef8 !important; font-weight: 700;
      border-top: 2px solid #d1fae5; color: #34d399;
    }}
    tr.row-aqs td:first-child {{ text-align: left; }}

    /* ── Dashboard panel ── */
    .pill-actions {{ display: flex; gap: 8px; margin-bottom: 12px; align-items: center; }}
    .pill-action-btn {{
      font-size: 12px; padding: 4px 14px; border-radius: 6px; cursor: pointer;
      border: 1px solid #d1d5db; background: #f9fafb; color: #374151; font-weight: 500;
    }}
    .pill-action-btn:hover {{ background: #e5e7eb; }}
    .brand-pills {{ display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 24px; }}
    .brand-pill {{
      padding: 5px 14px; border-radius: 20px; font-size: 12px; font-weight: 600;
      cursor: pointer; border: 2px solid rgba(0,0,0,.12); transition: opacity .15s;
      user-select: none;
    }}
    .brand-pill.inactive {{ opacity: .3; }}
    .dash-summary {{ display: flex; gap: 14px; flex-wrap: wrap; margin-bottom: 28px; }}
    .dash-card {{
      background: #f8fafc; border: 1px solid #e5e7eb; border-radius: 8px;
      padding: 14px 20px; min-width: 150px; flex: 1;
    }}
    .dash-card .dc-label {{ font-size: 11px; text-transform: uppercase;
                             letter-spacing: .05em; color: #6b7280; margin-bottom: 4px; }}
    .dash-card .dc-value {{ font-size: 24px; font-weight: 700; color: #111; }}
    .dash-card .dc-sub {{ font-size: 11px; color: #9ca3af; margin-top: 2px; }}
    .charts-row {{ display: flex; gap: 20px; flex-wrap: wrap; margin-bottom: 24px; }}
    .chart-box {{
      flex: 1; min-width: 280px; background: #fff; border: 1px solid #e5e7eb;
      border-radius: 8px; padding: 18px;
    }}
    .chart-box h3 {{ font-size: 12px; font-weight: 600; color: #374151;
                     text-transform: uppercase; letter-spacing: .04em; margin-bottom: 14px; }}
    .chart-box canvas {{ max-width: 100%; height: 240px !important; }}
    .dash-section-title {{
      font-size: 12px; font-weight: 600; color: #374151;
      text-transform: uppercase; letter-spacing: .04em; margin-bottom: 12px;
    }}
    .dash-table-wrap {{ overflow-x: auto; }}
    .dash-table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    .dash-table thead th {{
      text-align: right; padding: 8px 12px; border-bottom: 2px solid #e5e7eb;
      color: #666; font-weight: 600; font-size: 11px;
      text-transform: uppercase; letter-spacing: .05em; white-space: nowrap;
    }}
    .dash-table thead th:first-child {{ text-align: left; }}
    .dash-table tbody tr {{ border-bottom: 1px solid #f3f4f6; }}
    .dash-table tbody tr:hover {{ filter: brightness(.96); }}
    .dash-table tbody td {{ padding: 9px 12px; text-align: right; white-space: nowrap; }}
    .dash-table tbody td:first-child {{ text-align: left; font-weight: 600; }}

    .footer {{ margin-top: 32px; font-size: 11px; color: #9ca3af;
               border-top: 1px solid #e5e7eb; padding-top: 16px; line-height: 1.8; }}
    .footer code {{ background: #f3f4f6; padding: 1px 5px; border-radius: 3px; font-size: 11px; }}
  </style>
</head>
<body>

<h1>iOS AQS Consolidation Report</h1>
<p class="subtitle">
  Period: {date_range} &nbsp;|&nbsp; Fetched: {fetch_date}
</p>

{stale_warning_html}
{bq_status_note}

<div class="tabs" id="tabs">
  <div class="tab active" data-panel="aqs-table">AQS Table</div>
  <div class="tab" data-panel="dashboard">Dashboard</div>
  <div class="tab" data-panel="version-aqs">Version AQS</div>
</div>

<!-- ── Panel 1: AQS Table ── -->
<div class="panel active" id="panel-aqs-table">

<div class="table-wrap">
  <table>
    <thead>
      <tr>
        <th>Brand</th>
        <th class="num">Crash Free %</th>
        <th class="num">App Hangs %</th>
        <th class="num">App Size (MB)</th>
        <th class="num">ASTI (s)</th>
        <th class="num">STTI (s)</th>
        <th class="num">Frozen Frames %</th>
        <th class="num">Skipped Frames %</th>
        <th class="num">Riders</th>
      </tr>
    </thead>
    <tbody>
      {brand_rows_html}
      <tr class="row-avg">
        <td>Weighted AVG</td>
        <td>{avg_metrics['cfu']:.2f}%</td>
        <td>{avg_metrics['hang']:.2f}%</td>
        <td>{avg_metrics['app_size']}</td>
        <td>{avg_metrics['asti']}</td>
        <td>{avg_metrics['stti']}</td>
        <td>{avg_metrics['frozen']:.2f}%</td>
        <td>{avg_metrics['skipped']:.2f}%</td>
        <td></td>
      </tr>
      <tr class="row-aqs">
        <td>AQS Score</td>
        {aqs_cells_html}
        <td></td>
      </tr>
    </tbody>
  </table>
</div>

<div class="score-wrap">
  <div class="score-card">
    <div class="sc-label">Final AQS Score</div>
    <div class="sc-value">{final_aqs}</div>
    <div class="sc-sub">{date_range}</div>
  </div>
</div>

<div class="footer">
  <strong>Notes:</strong><br>
  &bull; <strong>Crash Free % / App Hangs %</strong> — computed from Sentry Discover API
    (<code>level:fatal handled:no !stack.package:*gpsmaster* !stack.package:*GPSTraveller* !user.id:*-*-*-*-* !issue:RIDER-APP-IOS-Z7 !issue:RIDER-APP-IOS-1BM !issue:RIDER-APP-IOS-3DVQ</code> / <code>!user.id:*-*-*-*-* app.in_foreground:True "*App hang* detected*"</code>)
    against BQ daily user counts for the period.<br>
  &bull; <strong>Frozen Frames % / Skipped Frames %</strong> — single aggregated value from Firebase Performance BQ,
    applied to all brands equally.<br>
{stale_footer_note}  &bull; <strong>Weighted AVG</strong> — SUMPRODUCT of brand values with rider-share weights.<br>
  &bull; <strong>AQS formula</strong> — <code>min(100, max(0, (((value - baseline) / (target - baseline)) * 50) + 50)) * weight%</code><br>
  &bull; BQ_AVAILABLE = <code>{bq_available_str}</code> &nbsp;|&nbsp;
    Fallback frozen=0.58, skipped=1.2 used when Firebase BQ is unavailable.<br>
  &bull; Rows marked with <span style="color:#9ca3af">*</span> used fallback CFU/hang values (BQ user count was zero).
</div>

</div><!-- /panel-aqs-table -->

<!-- ── Panel 2: Dashboard ── -->
<div class="panel" id="panel-dashboard">

  <div class="pill-actions">
    <button class="pill-action-btn" onclick="selectAllBrands()">Select All</button>
    <button class="pill-action-btn" onclick="clearAllBrands()">Clear All</button>
  </div>
  <div class="brand-pills" id="brand-pills"></div>

  <div class="dash-summary" id="dash-summary"></div>

  <div class="charts-row">
    <div class="chart-box">
      <h3>Crash Free % per Brand</h3>
      <canvas id="chart-cfu"></canvas>
    </div>
    <div class="chart-box">
      <h3>App Hangs % per Brand</h3>
      <canvas id="chart-hang"></canvas>
    </div>
  </div>

  <div class="charts-row">
    <div class="chart-box">
      <h3>Frozen &amp; Skipped Frames % per Brand</h3>
      <canvas id="chart-frames"></canvas>
    </div>
    <div class="chart-box">
      <h3>App Size (MB) per Brand</h3>
      <canvas id="chart-appsize"></canvas>
    </div>
  </div>

  <p class="dash-section-title">Brand Detail</p>
  <div class="dash-table-wrap">
    <table class="dash-table">
      <thead>
        <tr>
          <th>Brand</th>
          <th>Crash Free %</th>
          <th>App Hangs %</th>
          <th>App Size (MB)</th>
          <th>ASTI (s)</th>
          <th>STTI (s)</th>
          <th>Frozen Frames %</th>
          <th>Skipped Frames %</th>
          <th>Riders</th>
        </tr>
      </thead>
      <tbody id="dash-table-body"></tbody>
    </table>
  </div>

</div><!-- /panel-dashboard -->

<!-- ── Panel 3: Version AQS ── -->
<div class="panel" id="panel-version-aqs">
  {version_panel_html}
</div><!-- /panel-version-aqs -->

<script>
var BRAND_DATA = {brand_json};

// ── Tab switching ─────────────────────────────────────────────────────────────
document.getElementById('tabs').addEventListener('click', function(e) {{
  var t = e.target.closest('.tab');
  if (!t) return;
  document.querySelectorAll('.tab').forEach(function(x) {{ x.classList.remove('active'); }});
  document.querySelectorAll('.panel').forEach(function(x) {{ x.classList.remove('active'); }});
  t.classList.add('active');
  document.getElementById('panel-' + t.dataset.panel).classList.add('active');
}});

// ── Brand selection state ─────────────────────────────────────────────────────
var selected = BRAND_DATA.map(function() {{ return true; }});
var chartCfu, chartHang, chartFrames, chartAppsize;

function getActiveBrands() {{
  return BRAND_DATA.filter(function(b, i) {{ return selected[i]; }});
}}

function riderWeightedAvg(key) {{
  var active = getActiveBrands();
  if (!active.length) return 0;
  var totalRiders = 0, sum = 0;
  active.forEach(function(b) {{ sum += b[key] * b.riders; totalRiders += b.riders; }});
  return totalRiders ? sum / totalRiders : 0;
}}

// ── Summary cards ─────────────────────────────────────────────────────────────
function renderSummary() {{
  var active = getActiveBrands();
  var cards = [
    {{ key: 'cfu',      label: 'Crash Free %',     fmt: function(v) {{ return v.toFixed(2) + '%'; }} }},
    {{ key: 'hang',     label: 'App Hangs %',       fmt: function(v) {{ return v.toFixed(2) + '%'; }} }},
    {{ key: 'frozen',   label: 'Frozen Frames %',   fmt: function(v) {{ return v.toFixed(2) + '%'; }} }},
    {{ key: 'skipped',  label: 'Skipped Frames %',  fmt: function(v) {{ return v.toFixed(2) + '%'; }} }},
    {{ key: 'app_size', label: 'Avg App Size (MB)',  fmt: function(v) {{ return v.toFixed(1) + ' MB'; }} }},
  ];
  var sub = active.length + ' brand' + (active.length !== 1 ? 's' : '') + ' · rider-weighted';
  document.getElementById('dash-summary').innerHTML = cards.map(function(c) {{
    return '<div class="dash-card">'
      + '<div class="dc-label">' + c.label + '</div>'
      + '<div class="dc-value">' + c.fmt(riderWeightedAvg(c.key)) + '</div>'
      + '<div class="dc-sub">' + sub + '</div></div>';
  }}).join('');
}}

// ── Chart helpers ─────────────────────────────────────────────────────────────
function hexToRgba(hex, a) {{
  var r = parseInt(hex.slice(1,3),16), g = parseInt(hex.slice(3,5),16), b = parseInt(hex.slice(5,7),16);
  return 'rgba(' + r + ',' + g + ',' + b + ',' + a + ')';
}}
function darken(hex) {{
  var r = parseInt(hex.slice(1,3),16), g = parseInt(hex.slice(3,5),16), b = parseInt(hex.slice(5,7),16);
  return 'rgb(' + Math.max(0,r-60) + ',' + Math.max(0,g-60) + ',' + Math.max(0,b-60) + ')';
}}

function renderCharts() {{
  var active = getActiveBrands();
  var labels = active.map(function(b) {{ return b.name; }});
  var bgColors = active.map(function(b) {{ return hexToRgba(b.color, 0.88); }});
  var bdColors = active.map(function(b) {{ return darken(b.color); }});

  if (chartCfu) chartCfu.destroy();
  chartCfu = new Chart(document.getElementById('chart-cfu'), {{
    type: 'bar',
    data: {{ labels: labels, datasets: [{{
      data: active.map(function(b) {{ return b.cfu; }}),
      backgroundColor: bgColors, borderColor: bdColors, borderWidth: 1, borderRadius: 4
    }}] }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      plugins: {{ legend: {{ display: false }}, tooltip: {{ callbacks: {{ label: function(ctx) {{ return ctx.raw.toFixed(3) + '%'; }} }} }} }},
      scales: {{ y: {{ min: 99, max: 100.1, ticks: {{ callback: function(v) {{ return v + '%'; }} }} }} }},
    }},
  }});

  if (chartHang) chartHang.destroy();
  chartHang = new Chart(document.getElementById('chart-hang'), {{
    type: 'bar',
    data: {{ labels: labels, datasets: [{{
      data: active.map(function(b) {{ return b.hang; }}),
      backgroundColor: bgColors, borderColor: bdColors, borderWidth: 1, borderRadius: 4
    }}] }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      plugins: {{ legend: {{ display: false }}, tooltip: {{ callbacks: {{ label: function(ctx) {{ return ctx.raw.toFixed(3) + '%'; }} }} }} }},
      scales: {{ y: {{ min: 99, max: 100.1, ticks: {{ callback: function(v) {{ return v + '%'; }} }} }} }},
    }},
  }});

  if (chartFrames) chartFrames.destroy();
  chartFrames = new Chart(document.getElementById('chart-frames'), {{
    type: 'bar',
    data: {{
      labels: labels,
      datasets: [
        {{ label: 'Frozen %', data: active.map(function(b) {{ return b.frozen; }}),
           backgroundColor: 'rgba(99,102,241,.75)', borderColor: 'rgb(79,70,229)', borderWidth: 1, borderRadius: 4 }},
        {{ label: 'Skipped %', data: active.map(function(b) {{ return b.skipped; }}),
           backgroundColor: 'rgba(251,146,60,.75)', borderColor: 'rgb(234,88,12)', borderWidth: 1, borderRadius: 4 }},
      ],
    }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      plugins: {{ legend: {{ display: true, position: 'top' }},
                  tooltip: {{ callbacks: {{ label: function(ctx) {{ return ctx.dataset.label + ': ' + ctx.raw.toFixed(2) + '%'; }} }} }} }},
      scales: {{ y: {{ ticks: {{ callback: function(v) {{ return v + '%'; }} }} }} }},
    }},
  }});

  if (chartAppsize) chartAppsize.destroy();
  chartAppsize = new Chart(document.getElementById('chart-appsize'), {{
    type: 'bar',
    data: {{ labels: labels, datasets: [{{
      data: active.map(function(b) {{ return b.app_size; }}),
      backgroundColor: bgColors, borderColor: bdColors, borderWidth: 1, borderRadius: 4
    }}] }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      plugins: {{ legend: {{ display: false }}, tooltip: {{ callbacks: {{ label: function(ctx) {{ return ctx.raw + ' MB'; }} }} }} }},
      scales: {{ y: {{ min: 50, ticks: {{ callback: function(v) {{ return v + ' MB'; }} }} }} }},
    }},
  }});
}}

// ── Brand detail table ────────────────────────────────────────────────────────
function renderTable() {{
  var active = getActiveBrands();
  var tbody = document.getElementById('dash-table-body');
  if (!active.length) {{
    tbody.innerHTML = '<tr><td colspan="9" style="text-align:center;color:#9ca3af;padding:20px;">No brands selected</td></tr>';
    return;
  }}
  tbody.innerHTML = active.map(function(b) {{
    return '<tr style="background:' + b.color + ';">'
      + '<td>' + b.name + '</td>'
      + '<td>' + b.cfu.toFixed(2) + '%</td>'
      + '<td>' + b.hang.toFixed(2) + '%</td>'
      + '<td>' + b.app_size + '</td>'
      + '<td>' + b.asti + '</td>'
      + '<td>' + b.stti + '</td>'
      + '<td>' + b.frozen.toFixed(2) + '%</td>'
      + '<td>' + b.skipped.toFixed(2) + '%</td>'
      + '<td>' + b.riders.toLocaleString() + '</td>'
      + '</tr>';
  }}).join('');
}}

function renderAll() {{
  renderSummary();
  renderCharts();
  renderTable();
}}

// ── Brand pills ───────────────────────────────────────────────────────────────
function buildPills() {{
  document.getElementById('brand-pills').innerHTML = BRAND_DATA.map(function(b, i) {{
    return '<div class="brand-pill" id="pill-' + i + '" style="background:' + b.color + ';" onclick="toggleBrand(' + i + ')">' + b.name + '</div>';
  }}).join('');
}}

function toggleBrand(i) {{
  selected[i] = !selected[i];
  document.getElementById('pill-' + i).classList.toggle('inactive', !selected[i]);
  renderAll();
}}

function selectAllBrands() {{
  selected = BRAND_DATA.map(function() {{ return true; }});
  BRAND_DATA.forEach(function(b, i) {{ document.getElementById('pill-' + i).classList.remove('inactive'); }});
  renderAll();
}}

function clearAllBrands() {{
  selected = BRAND_DATA.map(function() {{ return false; }});
  BRAND_DATA.forEach(function(b, i) {{ document.getElementById('pill-' + i).classList.add('inactive'); }});
  renderAll();
}}

buildPills();
renderAll();
</script>

</body>
</html>"""

    os.makedirs(OUT_DIR, exist_ok=True)
    out_path = os.path.join(OUT_DIR, "index.html")
    with open(out_path, "w") as f:
        f.write(html)

    print(f"\nReport written to: {out_path}")


# ── Excel workbook generation ──────────────────────────────────────────────────

def add_consolidation_sheet(wb, firebase_data=None):
    """Second tab: per-brand summary.

    Crash free users (B) and App Hangs (C) are computed directly from the
    Dashboard tab's raw data (T-AF) using SUMIFS, so they always reflect
    whichever date range was fetched — no need to change B10 per brand.

    Frozen Frames (G) and Skipped Frames (H) come from firebase_data
    returned by fetch_firebase_frames(); falls back to static values
    when the BQ fetch is unavailable.
    """
    if firebase_data is None:
        firebase_data = {}
    ws = wb.create_sheet(title="Consolidation", index=0)

    headers = [
        "Brand", "Crash free users", "App Hangs", "App size",
        "ASTI", "STTI", "Frozen Frames", "Skipped Frames",
    ]
    # Column M (13) — Riders count per brand (order matches BRANDS list)
    RIDER_COL  = 13   # M
    # Column N (14) — fixed brand weights by rider share; order matches BRANDS list
    WEIGHT_COL = 14   # N

    # Static per-brand columns: (brand_name, app_size, asti, stti, frozen_fb, skipped_fb)
    # Brand name and numeric columns are derived from the module-level BRANDS list.
    # frozen_fb / skipped_fb are fallback values used when firebase_data is unavailable.
    static_cols = [
        (b["name"], b["app_size"], b["asti"], b["stti"], 0.58, 1.2)
        for b in BRANDS
    ]
    brands = [row[0] for row in static_cols]

    month_days = calendar.monthrange(START_DATE.year, START_DATE.month)[1]

    DATA_START = 2
    DATA_END   = DATA_START + len(brands) - 1   # row 11
    AVG_ROW    = DATA_END + 1                   # row 12
    AQS_ROW    = AVG_ROW + 1                    # row 13
    FINAL_ROW  = AQS_ROW + 2                    # row 15 (one blank gap)

    FINAL_BLUE = PatternFill("solid", fgColor="4472C4")

    # ── Header row ────────────────────────────────────────────────────────────
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(1, ci, value=h)
        cell.font = BOLD; cell.fill = DASH_HEADER_FILL
        cell.border = THIN_BOX
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # ── Brand data rows ───────────────────────────────────────────────────────
    for i, (brand, app_size, asti, stti, frozen_fb, skipped_fb) in enumerate(static_cols):
        r = DATA_START + i
        row_fill = EVEN_ROW_FILL if r % 2 == 0 else None

        def styled(col, value=None):
            cell = ws.cell(r, col, value=value)
            cell.border = THIN_BOX
            if row_fill:
                cell.fill = row_fill
            return cell

        styled(1, brand).font = BOLD

        # B: Crash free % — SUMIFS on Dashboard raw data (no B10 dependency)
        # Fallback uses M (riders) × month_days to approximate user-days when BQ is unavailable.
        styled(2).value = (
            f'=TRUNC(MAX(0,MIN(100,(1-SUMIFS(Dashboard!$X:$X,Dashboard!$Z:$Z,"*"&$A{r}&"*")'
            f'/IF(SUMIFS(Dashboard!$V:$V,Dashboard!$U:$U,"*"&$A{r}&"*")>0,'
            f'SUMIFS(Dashboard!$V:$V,Dashboard!$U:$U,"*"&$A{r}&"*"),'
            f'M{r}*{month_days}))*100)),2)'
        )
        # C: Hang free %
        styled(3).value = (
            f'=TRUNC(MAX(0,MIN(100,(1-SUMIFS(Dashboard!$AB:$AB,Dashboard!$AD:$AD,"*"&$A{r}&"*")'
            f'/IF(SUMIFS(Dashboard!$V:$V,Dashboard!$U:$U,"*"&$A{r}&"*")>0,'
            f'SUMIFS(Dashboard!$V:$V,Dashboard!$U:$U,"*"&$A{r}&"*"),'
            f'M{r}*{month_days}))*100)),2)'
        )
        styled(4, app_size)
        styled(5, asti)
        styled(6, stti)

        # G: Frozen Frames — single aggregated value from BQ, same for all brands
        styled(7, firebase_data.get("frozen",  frozen_fb))

        # H: Skipped Frames — single aggregated value from BQ, same for all brands
        styled(8, firebase_data.get("skipped", skipped_fb))

    # ── Column M: Riders count ────────────────────────────────────────────────
    rcol = get_column_letter(RIDER_COL)
    mhdr = ws.cell(1, RIDER_COL, value="Riders count")
    mhdr.font = BOLD; mhdr.fill = SETTINGS_FILL; mhdr.border = THIN_BOX
    mhdr.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, count in enumerate(RIDER_COUNTS):
        cell = ws.cell(DATA_START + i, RIDER_COL, value=count)
        cell.border = THIN_BOX
        if (DATA_START + i) % 2 == 0:
            cell.fill = EVEN_ROW_FILL

    # ── Column N: Weight Rider ID (fixed brand rider-share weights) ───────────
    wcol = get_column_letter(WEIGHT_COL)
    hdr = ws.cell(1, WEIGHT_COL, value="Weight Rider ID")
    hdr.font = BOLD; hdr.fill = SETTINGS_FILL; hdr.border = THIN_BOX
    hdr.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, w in enumerate(WEIGHTS):
        cell = ws.cell(DATA_START + i, WEIGHT_COL, value=w)
        cell.number_format = "0.00%"
        cell.border = THIN_BOX
        if (DATA_START + i) % 2 == 0:
            cell.fill = EVEN_ROW_FILL

    # ── AVG row (row 12) — weighted average using rider-share weights in col N ─
    ws.cell(AVG_ROW, 1, value="AVG").font = BOLD
    ws.cell(AVG_ROW, 1).fill = SUMMARY_LBL_FILL
    ws.cell(AVG_ROW, 1).border = THIN_BOX
    for ci in range(2, 9):
        cl = get_column_letter(ci)
        cell = ws.cell(AVG_ROW, ci)
        cell.value = (f"=TRUNC(SUMPRODUCT({cl}{DATA_START}:{cl}{DATA_END},"
                      f"{wcol}{DATA_START}:{wcol}{DATA_END}),2)")
        cell.font = BOLD; cell.fill = SUMMARY_LBL_FILL; cell.border = THIN_BOX
        cell.number_format = "0.00"   # plain number, no % sign

    # ── AQS row (row 13) ─────────────────────────────────────────────────────
    aqs_formulas = {
        2: f"=MIN(100,MAX(0,(((B{AVG_ROW}-99.7)/(99.9-99.7))*50)+50))*35%",
        3: f"=MIN(100,MAX(0,(((C{AVG_ROW}-99.7)/(99.9-99.7))*50)+50))*25%",
        4: f"=MIN(100,MAX(0,(((D{AVG_ROW}-75)/(60-75))*50)+50))*2%",
        5: f"=MIN(100,MAX(0,(((E{AVG_ROW}-4)/(2-4))*50)+50))*10%",
        6: f"=MIN(100,MAX(0,(((F{AVG_ROW}-1.5)/(0.5-1.5))*50)+50))*10%",
        7: f"=MIN(100,MAX(0,(((G{AVG_ROW}-3)/(1-3))*50)+50))*13%",
        8: f"=MIN(100,MAX(0,(((H{AVG_ROW}-2)/(1-2))*50)+50))*5%",
    }
    ws.cell(AQS_ROW, 1, value="AQS").font = BOLD
    ws.cell(AQS_ROW, 1).fill = SUMMARY_VAL_FILL
    ws.cell(AQS_ROW, 1).border = THIN_BOX
    for ci, val in aqs_formulas.items():
        cell = ws.cell(AQS_ROW, ci, value=val)
        cell.font = BOLD; cell.fill = SUMMARY_VAL_FILL; cell.border = THIN_BOX
        cell.number_format = "0.00"   # plain number, no % sign

    # M13: sum of M column data rows
    m13 = ws.cell(AQS_ROW, 13, value=f"=SUM(M{DATA_START}:M{DATA_END})")
    m13.font = BOLD; m13.fill = SUMMARY_VAL_FILL; m13.border = THIN_BOX

    # N13: sum of weights (validation — should equal 100%)
    n13 = ws.cell(AQS_ROW, WEIGHT_COL, value=f"=SUM({wcol}{DATA_START}:{wcol}{DATA_END})")
    n13.font = BOLD; n13.fill = SETTINGS_FILL; n13.border = THIN_BOX
    n13.number_format = "0.00%"

    # ── FINAL AQS ─────────────────────────────────────────────────────────────
    ws.cell(FINAL_ROW, 1, value="FINAL AQS =")
    ws.cell(FINAL_ROW, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(FINAL_ROW, 1).fill = FINAL_BLUE
    ws.cell(FINAL_ROW, 1).border = MEDIUM_BOX
    ws.cell(FINAL_ROW, 1).alignment = Alignment(horizontal="right")

    final_cell = ws.cell(FINAL_ROW, 2)
    final_cell.value = f"=ROUND(SUM(B{AQS_ROW}:H{AQS_ROW}),2)"
    final_cell.font = Font(bold=True, color="FFFFFF", size=13)
    final_cell.fill = FINAL_BLUE
    final_cell.border = MEDIUM_BOX
    final_cell.number_format = "0.00"

    # ── Manual-data note (current month only — prev month values are final) ──────
    if TODAY.day >= 3:
        NOTE_TEXT = ("⚠  App size, ASTI and STTI have the older values, not recently updated. "
                     "Update these columns manually before sharing. "
                     "For the accurate AQS score, update these values.")
        NOTE_FILL = PatternFill("solid", fgColor="FFEB9C")   # amber
        NOTE_FONT = Font(bold=True, color="9C5700")           # dark orange

        NOTE_ROW = FINAL_ROW + 2
        note_cell = ws.cell(NOTE_ROW, 1, value=NOTE_TEXT)
        note_cell.font = NOTE_FONT
        note_cell.fill = NOTE_FILL
        note_cell.border = MEDIUM_BOX
        note_cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=NOTE_ROW, start_column=1,
                       end_row=NOTE_ROW,   end_column=8)
        ws.row_dimensions[NOTE_ROW].height = 30

        MANUAL_COMMENT = ("Older values, not recently updated.\n"
                          "Update manually for accurate AQS score.")
        for col_letter in ("D", "E", "F"):
            ws[f"{col_letter}1"].comment = Comment(MANUAL_COMMENT, "Script")

    # ── Column widths & freeze ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 18
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 17
    ws.column_dimensions[rcol].width = 15  # M — Riders count
    ws.column_dimensions[wcol].width = 16  # N — Weight Rider ID
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "B2"


def add_release_version_sheets(wb, version_aqs_data):
    """Add two Excel tabs per pinned release version:

    1. 'v{ver}' — Dashboard-style: raw BQ/crash/hang data in columns T-AD,
       per-day formulas in A-N with brand selector, exactly like the main Dashboard.
    2. 'v{ver} Cons' — Consolidation-style: per-brand CFU%/hang-free% summary
       reading from the release dashboard tab, with weighted average and AQS.
    """
    if not version_aqs_data:
        return

    FINAL_BLUE  = PatternFill("solid", fgColor="4472C4")
    month_days  = calendar.monthrange(START_DATE.year, START_DATE.month)[1]
    FORMULA_END = END_DATE.day + 1
    DATA_END_D  = month_days + 1   # last date row in dashboard tab

    BRAND_DROPDOWN = ["foodora", "woowa", "foodpanda", "talabat", "efood", "glovo",
                      "hungerstation", "yemek", "foody", "pedidosya"]

    bq_cols    = ["dt", "appId", "user_count"]
    crash_cols = ["CRASH_USERS", "day", "Brand"]
    hang_cols  = ["HANG_USERS",  "day", "Brand"]

    for rel_data in version_aqs_data:
        ver        = rel_data["version"]
        brand_data = rel_data.get("brand_data", [])
        bq_rows_r  = rel_data.get("bq_rows_rel",    [])
        crash_rows_r = rel_data.get("crash_rows_rel", [])
        hang_rows_r  = rel_data.get("hang_rows_rel",  [])

        dash_name = f"v{ver}"[:31]
        cons_name = f"v{ver} Cons"[:31]

        # ── Tab 1: Dashboard-style ────────────────────────────────────────────────
        ws = wb.create_sheet(title=dash_name)

        # Raw data headers and values (columns T-AD — same layout as main Dashboard)
        for ci, name in enumerate(bq_cols, start=BQ_START_COL):
            cell = ws.cell(row=1, column=ci, value=name)
            cell.font = BOLD; cell.fill = BQ_HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        for ci, name in enumerate(crash_cols, start=CRASH_START_COL):
            cell = ws.cell(row=1, column=ci, value=name)
            cell.font = BOLD; cell.fill = CRASH_HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        for ci, name in enumerate(hang_cols, start=HANG_START_COL):
            cell = ws.cell(row=1, column=ci, value=name)
            cell.font = BOLD; cell.fill = HANG_HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        for ri, row in enumerate(bq_rows_r, start=2):
            for ci, key in enumerate(bq_cols, start=BQ_START_COL):
                ws.cell(row=ri, column=ci, value=row[key])
        for ri, row in enumerate(crash_rows_r, start=2):
            for ci, key in enumerate(crash_cols, start=CRASH_START_COL):
                ws.cell(row=ri, column=ci, value=row[key])
        for ri, row in enumerate(hang_rows_r, start=2):
            for ci, key in enumerate(hang_cols, start=HANG_START_COL):
                ws.cell(row=ri, column=ci, value=row[key])

        # Settings panel (A-B rows 1-10)
        settings = [
            ("CFU part in AQS",  60),
            ("base AQS score",   32.65),
            ("CFU minimum",      99.5),
            ("CFU maximum",      99.9),
            (None,               None),
            ("Hang minimum",     99.5),
            ("Hang maximum",     99.9),
            ("Hang part in AQS", 25),
            (None,               None),
            ("Brand",            "woowa"),
        ]
        for i, (label, value) in enumerate(settings, start=1):
            if label:
                a = ws.cell(row=i, column=1, value=label)
                a.font = BOLD; a.fill = SETTINGS_FILL
                b = ws.cell(row=i, column=2, value=value)
                b.fill = SETTINGS_FILL

        brand_dv = DataValidation(type="list",
                                  formula1='"' + ",".join(BRAND_DROPDOWN) + '"',
                                  allow_blank=False, showDropDown=False)
        ws.add_data_validation(brand_dv)
        brand_dv.add("B10")

        # Column headers (C-N)
        col_headers = {
            3:  "Date",
            4:  "Total users count (ios_User_Count)",
            5:  "Crashed users count (Query)",
            7:  "CFU %",
            11: "Hang user count (Query)",
            12: "Hang free %",
            14: "Projected AQS including hangs",
        }
        for col, header in col_headers.items():
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = BOLD; cell.fill = DASH_HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Date column C
        for day in range(1, month_days + 1):
            ws.cell(row=day + 1, column=3,
                    value=START_DATE.replace(day=day).strftime("%Y-%m-%d"))

        # Per-day formulas (same as main Dashboard, self-referencing T-AD of this sheet)
        for r in range(2, FORMULA_END + 1):
            ws.cell(r, 4).value  = f'=SUMIFS(V:V,U:U,"*"&$B$10&"*",T:T,C{r})'
            ws.cell(r, 5).value  = f'=SUMIFS(X:X,Z:Z,"*"&$B$10&"*",Y:Y,C{r})'
            ws.cell(r, 7).value  = (f'=ROUND(MAX(0,MIN(100,(1-E{r}/IF(D{r}>0,D{r},'
                                    f'SUMIFS(Consolidation!$M:$M,Consolidation!$A:$A,"*"&$B$10&"*")/{month_days}))*100)),2)')
            ws.cell(r, 8).value  = f'=MIN(1,(G{r}-$B$3)/($B$4-$B$3))*$B$1'
            ws.cell(r, 11).value = f'=SUMIFS(AB:AB,AD:AD,"*"&$B$10&"*",AC:AC,C{r})'
            ws.cell(r, 12).value = (f'=ROUND(MAX(0,MIN(100,100*(1-K{r}/IF(D{r}>0,D{r},'
                                    f'SUMIFS(Consolidation!$M:$M,Consolidation!$A:$A,"*"&$B$10&"*")/{month_days})))),2)')
            ws.cell(r, 13).value = f'=MAX(0,(L{r}-$B$6)/($B$7-$B$6)*$B$8)'
            ws.cell(r, 14).value = f'=$B$2+H{r}+M{r}'

        # Summary rows
        LBL = DATA_END_D + 1
        VAL = DATA_END_D + 2
        month_name = START_DATE.strftime("%B")
        summaries = [
            (4,  None,
             f"=AVERAGE(D2:D{FORMULA_END})"),
            (7,  "AVG CFU",
             f'=TRUNC(MAX(0,MIN(100,(1-SUMIFS(X:X,Z:Z,"*"&$B$10&"*")/IF(SUMIFS(V:V,U:U,"*"&$B$10&"*")>0,SUMIFS(V:V,U:U,"*"&$B$10&"*"),SUMIFS(Consolidation!$M:$M,Consolidation!$A:$A,"*"&$B$10&"*")*{month_days}))*100)),2)'),
            (8,  None,
             f"=AVERAGE(H2:H{FORMULA_END})"),
            (9,  f"{month_name} AQS score",
             f"=$B$2+H{VAL}"),
            (12, "AVG Hang-free",
             f'=TRUNC(MAX(0,MIN(100,(1-SUMIFS(AB:AB,AD:AD,"*"&$B$10&"*")/IF(SUMIFS(V:V,U:U,"*"&$B$10&"*")>0,SUMIFS(V:V,U:U,"*"&$B$10&"*"),SUMIFS(Consolidation!$M:$M,Consolidation!$A:$A,"*"&$B$10&"*")*{month_days}))*100)),2)'),
            (13, None,
             f"=MAX(0,(L{VAL}-$B$6)/($B$7-$B$6)*$B$8)"),
            (14, "Projected AQS including hangs",
             f"=$B$2+MAX(0,(G{VAL}-$B$3)/($B$4-$B$3)*($B$1-$B$8))+M{VAL}"),
        ]
        for col, label, formula in summaries:
            if label:
                lc = ws.cell(row=LBL, column=col, value=label)
                lc.font = BOLD
            ws.cell(row=VAL, column=col, value=formula)
        ws.cell(row=VAL, column=7).number_format  = "0.##"
        ws.cell(row=VAL, column=9).number_format  = "0.00"
        ws.cell(row=VAL, column=12).number_format = "0.##"
        ws.cell(row=VAL, column=14).number_format = "0.00"

        # Column widths & freeze
        for ci in range(BQ_START_COL, BQ_START_COL + 3):
            ws.column_dimensions[get_column_letter(ci)].width = 22
        ws.column_dimensions["W"].width = 4
        for ci in range(CRASH_START_COL, CRASH_START_COL + 3):
            ws.column_dimensions[get_column_letter(ci)].width = 22
        ws.column_dimensions["AA"].width = 4
        for ci in range(HANG_START_COL, HANG_START_COL + 3):
            ws.column_dimensions[get_column_letter(ci)].width = 22
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 28
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["K"].width = 24
        ws.column_dimensions["L"].width = 14
        ws.column_dimensions["N"].width = 30
        ws.row_dimensions[1].height = 45
        ws.freeze_panes = "C2"

        # ── Tab 2: Consolidation-style ────────────────────────────────────────────
        wc = wb.create_sheet(title=cons_name)

        if not brand_data:
            continue

        C_DATA_START = 2
        C_DATA_END   = C_DATA_START + len(brand_data) - 1
        C_AVG_ROW    = C_DATA_END + 1
        C_AQS_ROW    = C_AVG_ROW + 1
        C_FINAL_ROW  = C_AQS_ROW + 2
        RIDER_COL    = 9   # I
        WEIGHT_COL   = 10  # J

        # Headers
        cons_headers = [
            "Brand", "Crash free users", "App Hangs", "App size",
            "ASTI", "STTI", "Frozen Frames", "Skipped Frames",
        ]
        for ci, h in enumerate(cons_headers, start=1):
            cell = wc.cell(1, ci, value=h)
            cell.font = BOLD; cell.fill = DASH_HEADER_FILL; cell.border = THIN_BOX
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        ri_hdr = wc.cell(1, RIDER_COL, value="Riders count")
        ri_hdr.font = BOLD; ri_hdr.fill = SETTINGS_FILL; ri_hdr.border = THIN_BOX
        ri_hdr.alignment = Alignment(horizontal="center", wrap_text=True)

        wt_hdr = wc.cell(1, WEIGHT_COL, value="Weight Rider ID")
        wt_hdr.font = BOLD; wt_hdr.fill = SETTINGS_FILL; wt_hdr.border = THIN_BOX
        wt_hdr.alignment = Alignment(horizontal="center", wrap_text=True)

        # Per-brand rows — CFU%/hang-free% formulas read from the release dashboard tab
        for i, bd in enumerate(brand_data):
            r        = C_DATA_START + i
            row_fill = EVEN_ROW_FILL if r % 2 == 0 else None

            def _sc(col, value=None, fmt=None, bold=False, r=r, rf=row_fill):
                cell = wc.cell(r, col, value=value)
                cell.border = THIN_BOX
                if rf:
                    cell.fill = rf
                if fmt:
                    cell.number_format = fmt
                if bold:
                    cell.font = BOLD
                return cell

            _sc(1, bd["name"], bold=True)

            # CFU% — SUMIFS on release dashboard tab's raw crash/BQ columns
            _sc(2).value = (
                f"=TRUNC(MAX(0,MIN(100,(1-SUMIFS('{dash_name}'!$X:$X,'{dash_name}'!$Z:$Z,\"*\"&$A{r}&\"*\")"
                f"/IF(SUMIFS('{dash_name}'!$V:$V,'{dash_name}'!$U:$U,\"*\"&$A{r}&\"*\")>0,"
                f"SUMIFS('{dash_name}'!$V:$V,'{dash_name}'!$U:$U,\"*\"&$A{r}&\"*\"),"
                f"I{r}*{month_days}))*100)),2)"
            )
            # Hang-free%
            _sc(3).value = (
                f"=TRUNC(MAX(0,MIN(100,(1-SUMIFS('{dash_name}'!$AB:$AB,'{dash_name}'!$AD:$AD,\"*\"&$A{r}&\"*\")"
                f"/IF(SUMIFS('{dash_name}'!$V:$V,'{dash_name}'!$U:$U,\"*\"&$A{r}&\"*\")>0,"
                f"SUMIFS('{dash_name}'!$V:$V,'{dash_name}'!$U:$U,\"*\"&$A{r}&\"*\"),"
                f"I{r}*{month_days}))*100)),2)"
            )
            static = BRANDS[i] if i < len(BRANDS) else {}
            _sc(4, static.get("app_size", 75.0))
            _sc(5, static.get("asti",     3.93))
            _sc(6, rel_data.get("stti",   1.20))
            _sc(7, 0.58)   # frozen — fleet fallback (no per-release firebase data)
            _sc(8, 1.20)   # skipped

            riders_cell = _sc(RIDER_COL, bd["riders"])
            weight_cell = _sc(WEIGHT_COL, bd["weight"])
            weight_cell.number_format = "0.00%"

        # Weighted AVG row
        wcol = get_column_letter(WEIGHT_COL)
        wc.cell(C_AVG_ROW, 1, value="AVG").font = BOLD
        wc.cell(C_AVG_ROW, 1).fill = SUMMARY_LBL_FILL
        wc.cell(C_AVG_ROW, 1).border = THIN_BOX
        for ci in range(2, 9):
            cl   = get_column_letter(ci)
            cell = wc.cell(C_AVG_ROW, ci)
            cell.value = (f"=TRUNC(SUMPRODUCT({cl}{C_DATA_START}:{cl}{C_DATA_END},"
                          f"{wcol}{C_DATA_START}:{wcol}{C_DATA_END}),2)")
            cell.font = BOLD; cell.fill = SUMMARY_LBL_FILL; cell.border = THIN_BOX
            cell.number_format = "0.00"

        # AQS row
        aqs_formulas = {
            2: f"=MIN(100,MAX(0,(((B{C_AVG_ROW}-99.7)/(99.9-99.7))*50)+50))*35%",
            3: f"=MIN(100,MAX(0,(((C{C_AVG_ROW}-99.7)/(99.9-99.7))*50)+50))*25%",
            4: f"=MIN(100,MAX(0,(((D{C_AVG_ROW}-75)/(60-75))*50)+50))*2%",
            5: f"=MIN(100,MAX(0,(((E{C_AVG_ROW}-4)/(2-4))*50)+50))*10%",
            6: f"=MIN(100,MAX(0,(((F{C_AVG_ROW}-1.5)/(0.5-1.5))*50)+50))*10%",
            7: f"=MIN(100,MAX(0,(((G{C_AVG_ROW}-3)/(1-3))*50)+50))*13%",
            8: f"=MIN(100,MAX(0,(((H{C_AVG_ROW}-2)/(1-2))*50)+50))*5%",
        }
        wc.cell(C_AQS_ROW, 1, value="AQS").font = BOLD
        wc.cell(C_AQS_ROW, 1).fill = SUMMARY_VAL_FILL
        wc.cell(C_AQS_ROW, 1).border = THIN_BOX
        for ci, formula in aqs_formulas.items():
            cell = wc.cell(C_AQS_ROW, ci, value=formula)
            cell.font = BOLD; cell.fill = SUMMARY_VAL_FILL; cell.border = THIN_BOX
            cell.number_format = "0.00"
        n_aqs = wc.cell(C_AQS_ROW, WEIGHT_COL,
                        value=f"=SUM({wcol}{C_DATA_START}:{wcol}{C_DATA_END})")
        n_aqs.font = BOLD; n_aqs.fill = SETTINGS_FILL; n_aqs.border = THIN_BOX
        n_aqs.number_format = "0.00%"

        # Final AQS
        wc.cell(C_FINAL_ROW, 1, value="FINAL AQS =")
        wc.cell(C_FINAL_ROW, 1).font = Font(bold=True, color="FFFFFF")
        wc.cell(C_FINAL_ROW, 1).fill = FINAL_BLUE
        wc.cell(C_FINAL_ROW, 1).border = MEDIUM_BOX
        wc.cell(C_FINAL_ROW, 1).alignment = Alignment(horizontal="right")
        final_cell = wc.cell(C_FINAL_ROW, 2)
        final_cell.value = f"=ROUND(SUM(B{C_AQS_ROW}:H{C_AQS_ROW}),2)"
        final_cell.font  = Font(bold=True, color="FFFFFF", size=13)
        final_cell.fill  = FINAL_BLUE
        final_cell.border = MEDIUM_BOX
        final_cell.number_format = "0.00"

        # Column widths & freeze
        wc.column_dimensions["A"].width = 18
        for col in "BCDEFGH":
            wc.column_dimensions[col].width = 17
        wc.column_dimensions[get_column_letter(RIDER_COL)].width  = 15
        wc.column_dimensions[get_column_letter(WEIGHT_COL)].width = 16
        wc.row_dimensions[1].height = 40
        wc.freeze_panes = "B2"


def write_excel(bq_rows, hang_rows, crash_rows, path, firebase_data=None, version_aqs_data=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    bq_cols    = ["dt", "appId", "user_count"]
    crash_cols = ["CRASH_USERS", "day", "Brand"]
    hang_cols  = ["HANG_USERS",  "day", "Brand"]

    # ── Raw data (T-AD) ────────────────────────────────────────────────────────
    # BQ:    T(20)=dt, U(21)=appId, V(22)=user_count
    # Crash: X(24)=CRASH_USERS, Y(25)=day, Z(26)=Brand
    # Hang:  AB(28)=HANG_USERS, AC(29)=day, AD(30)=Brand

    for ci, name in enumerate(bq_cols, start=BQ_START_COL):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = BOLD; cell.fill = BQ_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for ci, name in enumerate(crash_cols, start=CRASH_START_COL):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = BOLD; cell.fill = CRASH_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for ci, name in enumerate(hang_cols, start=HANG_START_COL):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = BOLD; cell.fill = HANG_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(bq_rows, start=2):
        for ci, key in enumerate(bq_cols, start=BQ_START_COL):
            ws.cell(row=ri, column=ci, value=row[key])

    for ri, row in enumerate(crash_rows, start=2):
        for ci, key in enumerate(crash_cols, start=CRASH_START_COL):
            ws.cell(row=ri, column=ci, value=row[key])

    for ri, row in enumerate(hang_rows, start=2):
        for ci, key in enumerate(hang_cols, start=HANG_START_COL):
            ws.cell(row=ri, column=ci, value=row[key])

    for ci in range(BQ_START_COL, BQ_START_COL + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 22
    ws.column_dimensions["W"].width = 4   # spacer between BQ and Crash sections
    for ci in range(CRASH_START_COL, CRASH_START_COL + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 22
    ws.column_dimensions["AA"].width = 4  # spacer between Crash and Hang sections
    for ci in range(HANG_START_COL, HANG_START_COL + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 22

    # ── Dashboard (A-N) ────────────────────────────────────────────────────────
    # Settings: column A = label, column B = value, rows 1-10
    settings = [
        ("CFU part in AQS",  60),      # B1
        ("base AQS score",   32.65),   # B2
        ("CFU minimum",      99.5),    # B3
        ("CFU maximum",      99.9),    # B4
        (None,               None),
        ("Hang minimum",     99.5),    # B6
        ("Hang maximum",     99.9),    # B7
        ("Hang part in AQS", 25),      # B8
        (None,               None),
        ("Brand",            "woowa"), # B10 — change to filter by brand
    ]
    for i, (label, value) in enumerate(settings, start=1):
        if label:
            a = ws.cell(row=i, column=1, value=label)
            a.font = BOLD; a.fill = SETTINGS_FILL
            b = ws.cell(row=i, column=2, value=value)
            b.fill = SETTINGS_FILL

    # Brand dropdown in B10
    BRAND_DROPDOWN = ["foodora", "woowa", "foodpanda", "talabat", "efood", "glovo",
                      "hungerstation", "yemek", "foody", "pedidosya"]
    brand_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(BRAND_DROPDOWN) + '"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(brand_dv)
    brand_dv.add("B10")

    # Column headers (row 1, columns C-N)
    col_headers = {
        3:  "Date",
        4:  "Total users count (ios_User_Count)",
        5:  "Crashed users count (Query)",
        7:  "CFU %",
        11: "Hang user count (Query)",
        12: "Hang free %",
        14: "Projected AQS including hangs",
    }
    for col, header in col_headers.items():
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = BOLD; cell.fill = DASH_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # All calendar days of current month in column C (rows 2 to month_days+1)
    month_days = calendar.monthrange(START_DATE.year, START_DATE.month)[1]
    for day in range(1, month_days + 1):
        ws.cell(row=day + 1, column=3, value=START_DATE.replace(day=day).strftime("%Y-%m-%d"))

    DATA_END    = month_days + 1       # last date row (all calendar days filled in col C)
    FORMULA_END = END_DATE.day + 1    # last row with formulas — up to yesterday only

    # Per-day formulas only for rows with actual data (day 1 → yesterday)
    for r in range(2, FORMULA_END + 1):
        ws.cell(r, 4).value  = f'=SUMIFS(V:V,U:U,"*"&$B$10&"*",T:T,C{r})'
        ws.cell(r, 5).value  = f'=SUMIFS(X:X,Z:Z,"*"&$B$10&"*",Y:Y,C{r})'
        ws.cell(r, 7).value  = f'=ROUND(MAX(0,MIN(100,(1-E{r}/IF(D{r}>0,D{r},SUMIFS(Consolidation!$M:$M,Consolidation!$A:$A,"*"&$B$10&"*")/{month_days}))*100)),2)'
        ws.cell(r, 8).value  = f'=MIN(1,(G{r}-$B$3)/($B$4-$B$3))*$B$1'
        ws.cell(r, 11).value = f'=SUMIFS(AB:AB,AD:AD,"*"&$B$10&"*",AC:AC,C{r})'
        ws.cell(r, 12).value = f'=ROUND(MAX(0,MIN(100,100*(1-K{r}/IF(D{r}>0,D{r},SUMIFS(Consolidation!$M:$M,Consolidation!$A:$A,"*"&$B$10&"*")/{month_days})))),2)'
        ws.cell(r, 13).value = f'=MAX(0,(L{r}-$B$6)/($B$7-$B$6)*$B$8)'
        ws.cell(r, 14).value = f'=$B$2+H{r}+M{r}'

    # Summary label row and value row below all calendar dates
    LBL = DATA_END + 1
    VAL = DATA_END + 2
    month_name = START_DATE.strftime("%B")

    summaries = [
        (4,  None,                            f"=AVERAGE(D2:D{FORMULA_END})"),
        (7,  "AVG CFU",                       f'=TRUNC(MAX(0,MIN(100,(1-SUMIFS(X:X,Z:Z,"*"&$B$10&"*")/IF(SUMIFS(V:V,U:U,"*"&$B$10&"*")>0,SUMIFS(V:V,U:U,"*"&$B$10&"*"),SUMIFS(Consolidation!$M:$M,Consolidation!$A:$A,"*"&$B$10&"*")*{month_days}))*100)),2)'),
        (8,  None,                            f"=AVERAGE(H2:H{FORMULA_END})"),
        (9,  f"{month_name} AQS score",       f"=$B$2+H{VAL}"),
        (12, "AVG Hang-free",                 f'=TRUNC(MAX(0,MIN(100,(1-SUMIFS(AB:AB,AD:AD,"*"&$B$10&"*")/IF(SUMIFS(V:V,U:U,"*"&$B$10&"*")>0,SUMIFS(V:V,U:U,"*"&$B$10&"*"),SUMIFS(Consolidation!$M:$M,Consolidation!$A:$A,"*"&$B$10&"*")*{month_days}))*100)),2)'),
        (13, None,                            f"=MAX(0,(L{VAL}-$B$6)/($B$7-$B$6)*$B$8)"),
        (14, "Projected AQS including hangs", f"=$B$2+MAX(0,(G{VAL}-$B$3)/($B$4-$B$3)*($B$1-$B$8))+M{VAL}"),
    ]
    for col, label, formula in summaries:
        if label:
            lc = ws.cell(row=LBL, column=col, value=label)
            lc.font = BOLD
        ws.cell(row=VAL, column=col, value=formula)

    ws.cell(row=VAL, column=4).number_format  = "0"      # rider count — integer
    ws.cell(row=VAL, column=7).number_format  = "0.##"   # AVG CFU % — up to 2 decimal places
    ws.cell(row=VAL, column=8).number_format  = "0.00"   # AVG CFU AQS (hidden helper)
    ws.cell(row=VAL, column=9).number_format  = "0.00"   # month AQS score
    ws.cell(row=VAL, column=12).number_format = "0.##"   # AVG Hang-free % — up to 2 decimal places
    ws.cell(row=VAL, column=13).number_format = "0.00"   # Hang-free AQS (hidden helper)
    ws.cell(row=VAL, column=14).number_format = "0.00"   # Projected AQS

    # ── Borders, alternating rows & summary colors ───────────────────────────────

    # Settings region A1:B10
    for r in range(1, 11):
        for c in (1, 2):
            ws.cell(r, c).border = THIN_BOX

    # Dashboard header row C1:N1
    for c in range(3, 15):
        ws.cell(1, c).border = THIN_BOX

    # Data rows C-N: alternating row fill + thin borders
    for r in range(2, DATA_END + 1):
        for c in range(3, 15):
            cell = ws.cell(r, c)
            cell.border = THIN_BOX
            if r % 2 == 0:
                cell.fill = EVEN_ROW_FILL

    # Cols A-B for rows 11-DATA_END: thin border only (empty settings area)
    for r in range(11, DATA_END + 1):
        for c in (1, 2):
            ws.cell(r, c).border = THIN_BOX

    # Summary label row
    for c in range(4, 15):
        cell = ws.cell(LBL, c)
        cell.fill = SUMMARY_LBL_FILL
        cell.border = MEDIUM_BOX

    # Summary value row
    for c in range(4, 15):
        cell = ws.cell(VAL, c)
        cell.fill = SUMMARY_VAL_FILL
        cell.border = MEDIUM_BOX

    # Raw data section (T-AF): header + data borders with alternating rows
    max_raw = max(len(bq_rows), len(crash_rows), len(hang_rows)) + 1
    raw_cols = (list(range(BQ_START_COL, BQ_START_COL + 3)) +
                list(range(CRASH_START_COL, CRASH_START_COL + 3)) +
                list(range(HANG_START_COL, HANG_START_COL + 3)))
    for r in range(1, max_raw + 1):
        for c in raw_cols:
            cell = ws.cell(r, c)
            cell.border = THIN_BOX
            if r > 1 and r % 2 == 0:
                cell.fill = EVEN_ROW_FILL

    # Freeze the header row
    ws.freeze_panes = "C2"

    # Dashboard column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 4   # spacer
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].hidden = True   # CFU AQS Score — hidden helper for col N
    ws.column_dimensions["I"].width = 22
    ws.column_dimensions["J"].width = 4   # spacer
    ws.column_dimensions["K"].width = 24
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].hidden = True   # Hang-free AQS — hidden helper for col N
    ws.column_dimensions["N"].width = 30
    ws.row_dimensions[1].height = 45

    add_release_version_sheets(wb, version_aqs_data)
    add_consolidation_sheet(wb, firebase_data=firebase_data)

    # Make Consolidation the first visible tab when the workbook is opened
    wb.active = wb.worksheets[0]  # index 0 = Consolidation (inserted first)

    wb.save(path)
    print(f"Wrote {path}  ({len(bq_rows)} BQ rows, {len(crash_rows)} crash rows, {len(hang_rows)} hang rows)")


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    print(f"Fetching data for {BQ_START} → {BQ_END}\n")

    print("Fetching BigQuery iOS user counts...")
    bq_rows = fetch_bigquery()

    print("Fetching Sentry hangs (per-brand queries)...")
    raw_hang_rows = fetch_discover_per_brand(HANGS_QUERY)
    hang_rows = shape_rows(raw_hang_rows, "HANG_USERS")
    backfill_zero_rows(hang_rows, "HANG_USERS", HANGS_QUERY)

    print("Fetching Sentry crashes (per-brand queries)...")
    raw_crash_rows = fetch_discover_per_brand(CRASHES_QUERY, environment="production")
    crash_rows = shape_rows(raw_crash_rows, "CRASH_USERS")
    backfill_zero_rows(crash_rows, "CRASH_USERS", CRASHES_QUERY, environment="production")

    print("Fetching Firebase Performance frames...")
    firebase_data = fetch_firebase_frames()

    # Aggregate per-brand totals for HTML report (no extra network calls)
    bq_users_by_brand = {}
    for row in bq_rows:
        key = (row["appId"] or "").lower()
        bq_users_by_brand[key] = bq_users_by_brand.get(key, 0) + row["user_count"]

    crash_by_brand = aggregate_by_brand(crash_rows, "CRASH_USERS")
    hang_by_brand  = aggregate_by_brand(hang_rows,  "HANG_USERS")

    # ── Version-based AQS ─────────────────────────────────────────────────────
    # app_size: rider-weighted fleet average (no per-version data available).
    # ASTI / STTI: per-version measured values stored in PINNED_RELEASES.
    total_riders    = sum(b["riders"] for b in BRANDS)
    static_app_size = round(sum(b["app_size"] * b["riders"] for b in BRANDS) / total_riders, 2) if total_riders else 75.0

    print(f"\nFetching version-based AQS (brand-weighted, {len(PINNED_RELEASES)} releases)...")
    version_aqs_data = []
    for rel in PINNED_RELEASES:
        ver        = rel["version"]
        ver_asti   = rel["asti"]
        ver_stti   = rel["stti"]
        print(f"  v{ver} — BQ users by version + brand (last 30 days)...")
        ver_bq_rows, ver_bq_users_by_brand = fetch_bigquery_by_version(ver, bq_start=REL_BQ_START, bq_end=REL_BQ_END)
        ver_total_users = sum(ver_bq_users_by_brand.values())
        print(f"    → {ver_total_users:,} total users")

        print(f"  v{ver} — crashes per brand per day (Sentry, last 30 days)...")
        ver_raw_crash      = fetch_discover_per_brand_for_release(rel, CRASHES_QUERY, environment="production", start=REL_START, end=REL_END)
        ver_crash_rows     = shape_rows(ver_raw_crash, "CRASH_USERS")
        crash_by_brand_rel = aggregate_by_brand(ver_crash_rows, "CRASH_USERS")
        crash_users        = sum(crash_by_brand_rel.values())

        print(f"  v{ver} — hangs per brand per day (Sentry, last 30 days)...")
        ver_raw_hang      = fetch_discover_per_brand_for_release(rel, HANGS_QUERY, start=REL_START, end=REL_END)
        ver_hang_rows     = shape_rows(ver_raw_hang, "HANG_USERS")
        hang_by_brand_rel = aggregate_by_brand(ver_hang_rows, "HANG_USERS")
        hang_users        = sum(hang_by_brand_rel.values())
        if "frozen_frames" in rel and "skipped_frames" in rel:
            frozen  = rel["frozen_frames"]
            skipped = rel["skipped_frames"]
            print(f"  v{ver} — frames: using pinned values (frozen={frozen}, skipped={skipped})")
        else:
            print(f"  v{ver} — frames (BQ, last 30 days)...", end=" ", flush=True)
            frames  = fetch_firebase_frames_by_version(ver, bq_start=REL_BQ_START, bq_end=REL_BQ_END)
            print("done")
            frozen  = frames.get("frozen")
            skipped = frames.get("skipped")
        no_data = (crash_users == 0 and hang_users == 0 and frozen is None)

        # Brand-weighted CFU and hang using per-day raw BQ rows.
        # Days with 0 BQ users are excluded from both the user count and the crash/hang count.
        # Brands with no valid days or in excluded_brands are skipped.
        # Remaining weights are normalized so they sum to 1.0.
        excluded = {b.lower() for b in rel.get("excluded_brands", [])}

        # Build {bkey: {day: user_count}} from raw per-day BQ rows
        bq_brand_day = {}
        for row in ver_bq_rows:
            appid = (row.get("appId") or "").lower()
            day   = row.get("dt", "")
            count = int(row.get("user_count", 0) or 0)
            for brand in BRANDS:
                b = brand["bq_key"].lower()
                if b in appid:
                    if b not in bq_brand_day:
                        bq_brand_day[b] = {}
                    bq_brand_day[b][day] = bq_brand_day[b].get(day, 0) + count
                    break

        # Build {skey: {day: count}} from raw per-day Sentry rows
        crash_brand_day = {}
        for row in ver_crash_rows:
            skey = (row.get("_sentry_key") or "").lower()
            day  = row.get("day", "")
            cnt  = int(row.get("CRASH_USERS", 0) or 0)
            if skey:
                if skey not in crash_brand_day:
                    crash_brand_day[skey] = {}
                crash_brand_day[skey][day] = crash_brand_day[skey].get(day, 0) + cnt

        hang_brand_day = {}
        for row in ver_hang_rows:
            skey = (row.get("_sentry_key") or "").lower()
            day  = row.get("day", "")
            cnt  = int(row.get("HANG_USERS", 0) or 0)
            if skey:
                if skey not in hang_brand_day:
                    hang_brand_day[skey] = {}
                hang_brand_day[skey][day] = hang_brand_day[skey].get(day, 0) + cnt

        # Pre-compute per-brand totals using only days where BQ users > 0
        ver_brand_users = {}
        for brand in BRANDS:
            bkey    = brand["bq_key"].lower()
            bq_days = bq_brand_day.get(bkey, {})
            ver_brand_users[bkey] = sum(v for d, v in bq_days.items() if v > 0)

        active_weight_sum = sum(
            w for brand, w in zip(BRANDS, WEIGHTS)
            if brand["sentry_key"].lower() not in excluded
            and ver_brand_users.get(brand["bq_key"].lower(), 0) > 0
        ) or 1.0

        brand_data      = []
        brand_cfu_list  = []
        brand_hang_list = []
        for brand, weight in zip(BRANDS, WEIGHTS):
            skey_b  = brand["sentry_key"].lower()
            bkey    = brand["bq_key"].lower()
            bq_days = bq_brand_day.get(bkey, {})
            valid_days = {d for d, v in bq_days.items() if v > 0}
            users     = ver_brand_users[bkey]
            crashes_b = sum(crash_brand_day.get(skey_b, {}).get(d, 0) for d in valid_days)
            hangs_b   = sum(hang_brand_day.get(skey_b,  {}).get(d, 0) for d in valid_days)
            if skey_b in excluded:
                b_cfu = b_hang = None  # not rolled out — excluded from all calculations
            elif users > 0:
                b_cfu  = max(0.0, min(100.0, int((1 - crashes_b / users) * 10000) / 100))
                b_hang = max(0.0, min(100.0, int((1 - hangs_b   / users) * 10000) / 100))
                norm_weight = weight / active_weight_sum
                brand_cfu_list.append(b_cfu  * norm_weight)
                brand_hang_list.append(b_hang * norm_weight)
            else:
                b_cfu = b_hang = 100.0  # no users yet — excluded from weighted average
            brand_data.append({
                "name":     brand["name"],
                "weight":   weight,
                "riders":   brand["riders"],
                "users":    users,
                "crashes":  crashes_b,
                "hangs":    hangs_b,
                "cfu":      b_cfu,
                "hang":     b_hang,
                "excluded": skey_b in excluded,
            })

        cfu  = int(sum(brand_cfu_list)  * 100) / 100
        hang = int(sum(brand_hang_list) * 100) / 100

        metrics = {
            "cfu":      cfu,
            "hang":     hang,
            "app_size": static_app_size,
            "asti":     ver_asti,
            "stti":     ver_stti,
            "frozen":   frozen  if frozen  is not None else 0.58,
            "skipped":  skipped if skipped is not None else 1.2,
        }
        aqs_scores = {k: round(aqs_score(v, k), 4) for k, v in metrics.items()}
        final_aqs  = round(sum(aqs_scores.values()), 2)

        version_aqs_data.append({
            "version":     ver,
            "bq_start":    REL_BQ_START,
            "bq_end":      REL_BQ_END,
            "crash_users": crash_users,
            "hang_users":  hang_users,
            "cfu":         cfu,
            "hang":        hang,
            "app_size":    static_app_size,
            "asti":        ver_asti,
            "stti":        ver_stti,
            "frozen":      frozen  if frozen  is not None else 0.58,
            "skipped":     skipped if skipped is not None else 1.2,
            "aqs_scores":  aqs_scores,
            "final_aqs":      final_aqs,
            "no_data":        no_data,
            "brand_data":     brand_data,
            "bq_rows_rel":    ver_bq_rows,
            "crash_rows_rel": ver_crash_rows,
            "hang_rows_rel":  ver_hang_rows,
        })
        print(f"    → CFU={cfu:.2f}%  Hang={hang:.2f}%  ASTI={ver_asti}  STTI={ver_stti}  Frozen={frozen}%  Skipped={skipped}%  AQS={final_aqs}")

    os.makedirs("consolidation_report", exist_ok=True)
    write_excel(bq_rows, hang_rows, crash_rows, "consolidation_report/sentry_data.xlsx",
                firebase_data=firebase_data, version_aqs_data=version_aqs_data)

    # Fetch previous month data when showing dual-month tabs (days 3-6)
    prev_bq_users_by_brand = prev_crash_by_brand = prev_hang_by_brand = prev_firebase_data = None
    if SHOW_PREV_TAB:
        print(f"\nFetching previous month data ({PREV_BQ_START} → {PREV_BQ_END})...")
        prev_bq_rows = fetch_bigquery(PREV_BQ_START, PREV_BQ_END)

        print("Fetching Sentry hangs for previous month...")
        prev_raw_hang = fetch_discover_per_brand(HANGS_QUERY, start=PREV_START, end=PREV_END)
        prev_hang_shaped = shape_rows(prev_raw_hang, "HANG_USERS")

        print("Fetching Sentry crashes for previous month...")
        prev_raw_crash = fetch_discover_per_brand(CRASHES_QUERY, environment="production",
                                                   start=PREV_START, end=PREV_END)
        prev_crash_shaped = shape_rows(prev_raw_crash, "CRASH_USERS")

        prev_firebase_data = fetch_firebase_frames(PREV_BQ_START, PREV_BQ_END)

        prev_bq_users_by_brand = {}
        for row in prev_bq_rows:
            key = (row["appId"] or "").lower()
            prev_bq_users_by_brand[key] = prev_bq_users_by_brand.get(key, 0) + row["user_count"]

        prev_crash_by_brand = aggregate_by_brand(prev_crash_shaped, "CRASH_USERS")
        prev_hang_by_brand  = aggregate_by_brand(prev_hang_shaped,  "HANG_USERS")

    generate_html_report(
        bq_users_by_brand, crash_by_brand, hang_by_brand, firebase_data,
        prev_bq_users=prev_bq_users_by_brand,
        prev_crash=prev_crash_by_brand,
        prev_hang=prev_hang_by_brand,
        prev_firebase=prev_firebase_data,
        version_aqs_data=version_aqs_data,
    )


if __name__ == "__main__":
    main()
